"""Standalone acceptance: synthetic screen flow + xlsx round-trip (#3 and #4).

No demo seeding (removed). Seeds a small synthetic universe + price snapshot and
the bundled FactSet dictionary directly via settings_store, then runs the screen
and exercises the formula generator.
"""
import io
import json
import os
import sys
import tempfile
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

# isolated DB
tmp = tempfile.mkdtemp()
os.environ["DB_PATH"] = os.path.join(tmp, "acc.db")
os.environ["APP_SECRET"] = "acc-secret"

from app.db import init_db
from app import settings_store as ss, formula_gen as fg
from app.web.common import run_active_screen

init_db()


def _synthetic():
    rng = np.random.default_rng(1)
    dates = pd.bdate_range("2024-01-01", periods=120)
    specs = {"AAA": ("Banks", -0.20), "BBB": ("Banks", 0.0), "CCC": ("Banks", 0.0),
             "DDD": ("Tech", 0.25), "EEE": ("Tech", 0.0)}
    rows, uni = [], []
    for t, (sub, shock) in specs.items():
        rets = rng.normal(0.0002, 0.012, 120)
        if shock:
            rets[-7:] += (1 + shock) ** (1 / 7) - 1
        prices = 100.0 * np.cumprod(1 + rets)
        for d, p in zip(dates, prices):
            rows.append({"ticker": t, "date": d.date().isoformat(),
                         "close": float(p), "volume": 1_000_000})
        uni.append({"ticker": t, "name": t, "sector": "X", "sub_industry": sub,
                    "index_weight": 1.0, "adv_usd_20d": 50_000_000, "below_floor": False})
    return pd.DataFrame(uni), pd.DataFrame(rows)


uni_df, px_df = _synthetic()
ss.add_universe(uni_df.to_csv(index=False), filename="acc_universe.csv", make_active=True)
ss.add_snapshot(px_df.to_csv(index=False), filename="acc_prices.csv", make_active=True)

# bundled FactSet dictionary (kept for formula generation)
dict_path = Path(ROOT) / "sample_data" / "dictionary.json"
ss.add_dictionary(dict_path.read_text(), filename="dictionary.json", make_active=True)

# ---- Acceptance #3: run screen, assert non-empty + |z|-ranked ----
res = run_active_screen()
assert not res.get("_empty"), "screen returned empty"
os_n = len(res["oversold"]); ob_n = len(res["overbought"])
print(f"[#3] oversold-reversion={os_n}  overbought-fade={ob_n}")
assert os_n > 0, "oversold list empty"
assert ob_n > 0, "overbought list empty"
abs_z = res["master"]["abs_z"].dropna().tolist()
assert abs_z == sorted(abs_z, reverse=True), "master not ranked by |z|"
assert (res["master"]["dislocation_type"] == "IDIOSYNCRATIC").any(), "no idiosyncratic example"
print("[#3] both lists non-empty, master ranked by |z|, idiosyncratic example present  ✓")

# ---- Acceptance #4: formula generator produces a valid downloadable xlsx ----
d = ss.get_active_dictionary()
tickers = ["BABA-CN", "PDD-CN"]
for method in ("A", "B"):
    data = fg.build_formula_workbook(tickers, d["data"], method=method, lookback=50)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Instructions" in wb.sheetnames
    assert "BABA-CN" in wb.sheetnames
    print(f"[#4] Method {method} xlsx opens; sheets={wb.sheetnames[:4]}...  ✓")

print("\nALL STANDALONE ACCEPTANCE CHECKS PASSED")
