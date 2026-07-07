"""IMPROVEMENT 2: earnings + ex-dividend FactSet event-date pulls.

Covers the three layers:
  1. formula_gen  -> emits the two event-date =FDS formulas (with doubled-quote
     escaping for the ex-dividend template) when events are enabled; omits them
     when disabled.
  2. data_ingest  -> decodes a YYYYMMDD ex-div int (e.g. 20260526) and a next-
     earnings date to real Timestamps; leaves things alone when the columns are
     absent (backward-compatible).
  3. screen_engine -> sets event_flag True + populates event_date when a pulled
     ex-div/earnings date falls inside the event window; False when absent.
  4. dictionary    -> ships an ``ex_dividend_date`` entry whose template stores
     the FCA_EVENT_DATE args with single double-quotes (doubling at emit time).
"""
import io
import json
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from app import data_ingest as di
from app import formula_gen as fg
from app import screen_engine as se

ROOT = Path(__file__).resolve().parent.parent
DICT = json.loads((ROOT / "sample_data" / "dictionary.json").read_text())

# Verified exact emitted strings (doubled double-quotes round-trip in Excel).
# Ex-dividend keeps the =FDS FCA_EVENT_DATE pull UNCHANGED. Earnings switched to
# the LIVE =FDSLIVE function with the RTP_ fields (no nested quotes, no date args).
EX_DIV_LITERAL = '=FDS("9988-HK","FCA_EVENT_DATE(0,""CASH_DIST"",""EXDATE"",""YYYYMMDD"")")'
EX_DIV_CELLREF = '=FDS(A2,"FCA_EVENT_DATE(0,""CASH_DIST"",""EXDATE"",""YYYYMMDD"")")'
EARN_LITERAL = '=FDSLIVE("9988-HK","RTP_EARNINGS_RELEASE_DATE")'
EARN_CELLREF = '=FDSLIVE(A2,"RTP_EARNINGS_RELEASE_DATE")'
STATUS_LITERAL = '=FDSLIVE("9988-HK","RTP_EARNINGS_RELEASE_STATUS")'
STATUS_CELLREF = '=FDSLIVE(A2,"RTP_EARNINGS_RELEASE_STATUS")'


# --- Layer 1: formula generator ------------------------------------------

def test_single_cell_event_formulas_exact_escaping():
    # Ex-dividend =FDS pull is UNCHANGED (doubled-quote escaping).
    assert fg.ex_dividend_formula("9988-HK", DICT) == EX_DIV_LITERAL
    assert fg.ex_dividend_formula("A2", DICT) == EX_DIV_CELLREF
    # Earnings date + status now emit =FDSLIVE with the RTP_ fields.
    assert fg.earnings_date_formula("9988-HK", DICT) == EARN_LITERAL
    assert fg.earnings_date_formula("A2", DICT) == EARN_CELLREF
    assert fg.earnings_status_formula("9988-HK", DICT) == STATUS_LITERAL
    assert fg.earnings_status_formula("A2", DICT) == STATUS_CELLREF


def test_fdslive_helper_literal_and_cell_forms():
    assert fg.fdslive_formula("9988-HK", "RTP_EARNINGS_RELEASE_DATE") == EARN_LITERAL
    assert fg.fdslive_formula("A2", "RTP_EARNINGS_RELEASE_DATE") == EARN_CELLREF
    assert fg.fdslive_formula("$A$2", "RTP_EARNINGS_RELEASE_STATUS") == \
        '=FDSLIVE($A$2,"RTP_EARNINGS_RELEASE_STATUS")'


def _headers(data, sheet=None):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    if sheet is None:
        sheet = [s for s in wb.sheetnames if s != "Instructions"][0]
    ws = wb[sheet]
    return [c.value for c in ws[1]], ws, wb


def _all_formula_text(data):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    out = []
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for c in row:
                if c.value is not None:
                    out.append(str(c.value))
    return "\n".join(out)


def test_workbook_emits_both_event_formulas_when_enabled():
    for layout in ("spill", "stacked", "per_ticker"):
        data = fg.build_formula_workbook(["9988-HK"], DICT, method="A",
                                         layout=layout, lookback=20,
                                         include_events=True)
        hdrs, _ws, _wb = _headers(data)
        assert "earnings_date" in hdrs, layout
        assert "earnings_status" in hdrs, layout
        assert "ex_dividend_date" in hdrs, layout
        text = _all_formula_text(data)
        # doubled-quote ex-div =FDS template present (UNCHANGED) ...
        assert 'FCA_EVENT_DATE(0,""CASH_DIST"",""EXDATE"",""YYYYMMDD"")' in text, layout
        # ... and BOTH RTP earnings pulls via the LIVE =FDSLIVE function.
        assert '=FDSLIVE' in text, layout
        assert 'RTP_EARNINGS_RELEASE_DATE' in text, layout
        assert 'RTP_EARNINGS_RELEASE_STATUS' in text, layout
        # the OLD earnings field must be gone
        assert "FE_REP_DT_NEXT" not in text, layout


def test_method_b_emits_both_event_formulas_when_enabled():
    data = fg.build_formula_workbook(["9988-HK"], DICT, method="B", lookback=20,
                                     include_events=True)
    text = _all_formula_text(data)
    assert 'FCA_EVENT_DATE(0,""CASH_DIST"",""EXDATE"",""YYYYMMDD"")' in text
    assert '=FDSLIVE(' in text
    assert 'RTP_EARNINGS_RELEASE_DATE' in text
    assert 'RTP_EARNINGS_RELEASE_STATUS' in text
    assert "FE_REP_DT_NEXT" not in text


def test_workbook_omits_event_formulas_when_disabled():
    for layout in ("spill", "stacked", "per_ticker"):
        data = fg.build_formula_workbook(["9988-HK"], DICT, method="A",
                                         layout=layout, lookback=20,
                                         include_events=False)
        hdrs, _ws, _wb = _headers(data)
        assert "earnings_date" not in hdrs, layout
        assert "earnings_status" not in hdrs, layout
        assert "ex_dividend_date" not in hdrs, layout
        text = _all_formula_text(data)
        assert "FCA_EVENT_DATE" not in text, layout
        assert "FDSLIVE" not in text, layout
        assert "RTP_EARNINGS" not in text, layout


def test_spill_price_vol_columns_untouched_by_events():
    # Event columns append AFTER the close/volume block so the spill-activation
    # macro's B/C/D (date/close/volume) references are preserved.
    data = fg.build_formula_workbook(["9988-HK"], DICT, method="A",
                                     layout="spill", lookback=109,
                                     include_events=True)
    _hdrs, ws, _wb = _headers(data, sheet="9988-HK")
    def _af(v):
        return str(v).replace("_xlfn._xlws.", "").replace("_xlfn.", "")
    assert _af(ws.cell(row=2, column=3).value) == '=FDS(A2,"P_PRICE(0,-109D,D)")'
    assert _af(ws.cell(row=2, column=4).value) == '=FDS(A2,"P_VOLUME_DAY(0,-109D,D)")'


# --- Layer 2: data_ingest decode -----------------------------------------

def test_decode_event_date_yyyymmdd_int():
    assert di._decode_event_date(20260526) == pd.Timestamp("2026-05-26")
    assert di._decode_event_date("20260526") == pd.Timestamp("2026-05-26")
    assert di._decode_event_date(20260526.0) == pd.Timestamp("2026-05-26")


def test_decode_event_date_datestring_and_serial():
    assert di._decode_event_date("2026-05-26") == pd.Timestamp("2026-05-26")
    # Excel/FactSet-Julian serial (1899-12-30 origin) for 2026-05-26
    serial = (pd.Timestamp("2026-05-26") - pd.Timestamp("1899-12-30")).days
    assert 20000 <= serial <= 80000
    assert di._decode_event_date(serial) == pd.Timestamp("2026-05-26")


def test_decode_event_date_blank_and_errors_none():
    for v in (None, "", "  ", "nan", "#N/A", "#ERR", "@NA", float("nan")):
        assert di._decode_event_date(v) is None


# --- Layer 3: screen_engine event flagging --------------------------------

def _uni_prices():
    rng = np.random.default_rng(3)
    dates = pd.bdate_range("2024-01-01", periods=120)
    rows, uni = [], []
    for t, sub, shock in (("AAA", "Banks", -0.20), ("BBB", "Banks", 0.0),
                          ("CCC", "Banks", 0.0), ("DDD", "Tech", 0.25),
                          ("EEE", "Tech", 0.0)):
        rets = rng.normal(0.0002, 0.012, 120)
        if shock:
            rets[-7:] += (1 + shock) ** (1 / 7) - 1
        px = 100.0 * np.cumprod(1 + rets)
        for d, p in zip(dates, px):
            rows.append({"ticker": t, "date": d, "close": float(p),
                         "volume": 1_000_000})
        uni.append({"ticker": t, "name": t, "sector": "X", "sub_industry": sub,
                    "index_weight": 1.0, "adv_usd_20d": 50_000_000,
                    "below_floor": False})
    return pd.DataFrame(uni), pd.DataFrame(rows), dates.max()


def test_select_event_date_in_and_out_of_window():
    asof = pd.Timestamp("2026-06-14")
    # ex-div 3 days AGO -> in window (ex-div is ±window)
    ed, inw = se._select_event_date(None, "20260611", asof, 7)
    assert ed == pd.Timestamp("2026-06-11") and inw is True
    # earnings 3 days AHEAD -> in window (earnings is 0..window ahead)
    ed, inw = se._select_event_date("20260617", None, asof, 7)
    assert ed == pd.Timestamp("2026-06-17") and inw is True
    # earnings 3 days AGO -> NOT in window (past earnings not upcoming), but the
    # nearest date is still surfaced with flag False
    ed, inw = se._select_event_date("20260611", None, asof, 7)
    assert ed == pd.Timestamp("2026-06-11") and inw is False
    # nothing parseable -> (None, False)
    assert se._select_event_date(None, None, asof, 7) == (None, False)


def test_run_screen_event_flag_from_pulled_exdiv_in_window():
    uni, prices, asof = _uni_prices()
    prices = prices.copy()
    # attach an in-window ex-div (2 days after asof) to AAA only
    exdiv = int((pd.Timestamp(asof) + pd.Timedelta(days=2)).strftime("%Y%m%d"))
    prices["ex_dividend_date"] = np.nan
    prices.loc[prices["ticker"] == "AAA", "ex_dividend_date"] = exdiv
    res = se.run_screen(prices, uni, dict(se.DEFAULT_PARAMS, min_bars=60))
    m = res["master"].set_index("ticker")
    assert bool(m.loc["AAA", "event_flag"]) is True
    assert pd.to_datetime(m.loc["AAA", "event_date"]) == \
        pd.Timestamp(asof) + pd.Timedelta(days=2)
    # a ticker with no event date stays unflagged
    assert bool(m.loc["BBB", "event_flag"]) is False
    assert res["meta"]["event_data_loaded"] is True


def test_run_screen_no_event_columns_backward_compatible():
    uni, prices, _asof = _uni_prices()
    res = se.run_screen(prices, uni, dict(se.DEFAULT_PARAMS, min_bars=60))
    m = res["master"].set_index("ticker")
    assert bool(m.loc["AAA", "event_flag"]) is False
    assert res["meta"]["event_data_loaded"] is False


def test_run_screen_earnings_outside_window_not_flagged():
    uni, prices, asof = _uni_prices()
    prices = prices.copy()
    far = int((pd.Timestamp(asof) + pd.Timedelta(days=60)).strftime("%Y%m%d"))
    prices["earnings_date"] = np.nan
    prices.loc[prices["ticker"] == "AAA", "earnings_date"] = far
    res = se.run_screen(prices, uni, dict(se.DEFAULT_PARAMS, min_bars=60))
    m = res["master"].set_index("ticker")
    # nearest date surfaced, but far outside the 7-day window -> flag False
    assert bool(m.loc["AAA", "event_flag"]) is False
    assert pd.to_datetime(m.loc["AAA", "event_date"]) == \
        pd.Timestamp(asof) + pd.Timedelta(days=60)


# --- Layer 4: dictionary --------------------------------------------------

def test_dictionary_has_ex_dividend_entry_with_single_quote_template():
    f = DICT["formulas"]
    assert "ex_dividend_date" in f
    tmpl = f["ex_dividend_date"]["fql_template"]
    # stored with SINGLE double-quotes (doubling happens at emit time)
    assert tmpl == 'FCA_EVENT_DATE(0,"CASH_DIST","EXDATE","YYYYMMDD")'
    assert '""' not in tmpl
    assert f["ex_dividend_date"]["family"] == "corporate_actions"


def test_dictionary_has_next_earnings_entry():
    f = DICT["formulas"]
    assert "next_earnings" in f
    # Earnings now resolves to the LIVE RTP_ field (pulled via =FDSLIVE).
    assert f["next_earnings"]["fql_template"] == "RTP_EARNINGS_RELEASE_DATE"
    assert f["next_earnings"].get("fds_compatible") is False


def test_dictionary_has_earnings_release_status_entry():
    f = DICT["formulas"]
    assert "earnings_release_status" in f
    assert f["earnings_release_status"]["fql_template"] == "RTP_EARNINGS_RELEASE_STATUS"
    assert f["earnings_release_status"].get("fds_compatible") is False


def test_decode_earnings_rtp_int_and_status_text():
    # RTP earnings date returns a YYYYMMDD int (e.g. 20260831) -> real date.
    assert di._decode_event_date(20260831) == pd.Timestamp("2026-08-31")
    assert di._decode_event_date("20260831") == pd.Timestamp("2026-08-31")
    # RTP earnings status is free text carried through as-is.
    assert di._clean_status_value("Projected") == "Projected"
    assert di._clean_status_value("  Confirmed ") == "Confirmed"
    for v in (None, "", "nan", "#N/A", float("nan")):
        assert di._clean_status_value(v) is None
