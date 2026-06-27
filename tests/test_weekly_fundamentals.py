"""Phase D-2: fundamentals (valuation + EPS revisions via FE_ESTIMATE) +
sector-vs-stock-specific attribution.

Covers: template emits the exact LOCKED FE_ESTIMATE / FG strings (and a lean
template omits them); ingest round-trips fundamental cells incl. blanks/#N/A ->
None and FactSet classification text capture, and a 2-column (price/volume-only) template still
parses; valuation fwd_pe = price / FY1 EPS (None when EPS<=0/missing); EPS
momentum (revision abs/pct/dir, zero-base guard, dispersion); attribution
(stock-specific / sector-driven / idiosyncratic-solo) with leave-one-out that
excludes self; and the note renders a fundamentals/attribution table when data
is present and omits it otherwise.
"""
import io

import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from app.weekly import attribution as A
from app.weekly import ingest as wing
from app.weekly import metrics as M
from app.weekly import note as N
from app.weekly import template_gen as wtpl


# ---------------------------------------------------------------------------
# Task 1 — template emits the exact locked FE_ESTIMATE / FG_FACTSET formulas
# ---------------------------------------------------------------------------
def test_fundamental_formulas_exact_locked_strings():
    ff = wtpl.fundamental_formulas("A2")
    assert ff["fy1_eps_mean"] == '=FDS(A2,"FE_ESTIMATE(EPS,MEAN,ANN_ROLL,+1,NOW,,,\'\')")'
    assert ff["fy2_eps_mean"] == '=FDS(A2,"FE_ESTIMATE(EPS,MEAN,ANN_ROLL,+2,NOW,,,\'\')")'
    assert ff["fy1_eps_mean_4wk_ago"] == '=FDS(A2,"FE_ESTIMATE(EPS,MEAN,ANN_ROLL,+1,-20D,,,\'\')")'
    assert ff["fy1_eps_stddev"] == '=FDS(A2,"FE_ESTIMATE(EPS,STDDEV,ANN_ROLL,+1,NOW,,,\'\')")'
    assert ff["fy1_eps_num_est"] == '=FDS(A2,"FE_ESTIMATE(EPS,NEST,ANN_ROLL,+1,NOW,,,\'\')")'
    assert ff["factset_sector"] == '=FDS(A2,"FG_FACTSET_SECTOR")'
    assert ff["factset_industry"] == '=FDS(A2,"FG_FACTSET_IND")'
    # No recommendation mark / FE_UP / FE_DOWN anywhere.
    blob = " ".join(ff.values())
    assert "FE_UP" not in blob and "FE_DOWN" not in blob and "RECOMM" not in blob.upper()


def test_template_with_fundamentals_writes_block_as_text():
    b = wtpl.build_weekly_template(["0700-HK"], include_fundamentals=True)
    ws = load_workbook(io.BytesIO(b))["0700-HK"]
    # G2..G8 carry the 7 fundamental formulas as TEXT (data_type 's'), label in F.
    assert ws.cell(row=2, column=wtpl.FUND_FORMULA_COL).value == \
        '=FDS(A2,"FE_ESTIMATE(EPS,MEAN,ANN_ROLL,+1,NOW,,,\'\')")'
    assert ws.cell(row=2, column=wtpl.FUND_FORMULA_COL).data_type == "s"
    assert ws.cell(row=7, column=wtpl.FUND_FORMULA_COL).value == '=FDS(A2,"FG_FACTSET_SECTOR")'
    assert ws.cell(row=8, column=wtpl.FUND_FORMULA_COL).value == '=FDS(A2,"FG_FACTSET_IND")'
    assert ws.cell(row=2, column=wtpl.FUND_LABEL_COL).value == "FY1 EPS mean"


def test_lean_template_omits_fundamentals_block():
    b = wtpl.build_weekly_template(["0700-HK"], include_fundamentals=False)
    ws = load_workbook(io.BytesIO(b))["0700-HK"]
    # No fundamental label/formula written.
    assert ws.cell(row=2, column=wtpl.FUND_LABEL_COL).value in (None, "")
    assert ws.cell(row=2, column=wtpl.FUND_FORMULA_COL).value in (None, "")


# ---------------------------------------------------------------------------
# Task 3 — ingest round-trip: populate fundamentals, blanks/#N/A -> None, FactSet
# ---------------------------------------------------------------------------
def _julian(dates):
    origin = pd.Timestamp("1899-12-30")
    return [(pd.Timestamp(d) - origin).days for d in dates]


def _populate(tickers, fundamentals=None, include_fundamentals=True,
              n_bars=130, last_date="2026-06-26"):
    """Build the template (optionally with the fundamentals block), fill the
    price/volume series the way FactSet would, and fill the fundamental VALUE
    column (G) with provided per-ticker values."""
    fundamentals = fundamentals or {}
    b = wtpl.build_weekly_template(list(tickers),
                                   include_fundamentals=include_fundamentals)
    wb = load_workbook(io.BytesIO(b))
    dates = pd.bdate_range(end=last_date, periods=n_bars)
    jul = _julian(dates)
    for name in wb.sheetnames:
        if name in ("Instructions", "Manifest"):
            continue
        ws = wb[name]
        if name == "HSI":
            closes = list(np.linspace(20000, 21000, n_bars))
            for i, (jd, cl) in enumerate(zip(jul, closes)):
                ws.cell(row=2 + i, column=2, value=jd)
                ws.cell(row=2 + i, column=3, value=float(cl))
            continue
        closes = list(np.linspace(100, 120, n_bars))
        for i, (jd, cl) in enumerate(zip(jul, closes)):
            ws.cell(row=2 + i, column=2, value=jd)
            ws.cell(row=2 + i, column=3, value=float(cl))
            ws.cell(row=2 + i, column=4, value=1_000_000.0)
        # Fill the fundamental VALUE column (G) — emulate FactSet returning a
        # static value into each formula cell.
        if include_fundamentals and name in fundamentals:
            fnd = fundamentals[name]
            for i, (key, _label, _expr) in enumerate(wtpl.FUNDAMENTAL_FIELDS):
                r = 2 + i
                if key in fnd:
                    ws.cell(row=r, column=wtpl.FUND_FORMULA_COL, value=fnd[key])
    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()


def test_ingest_round_trips_fundamentals_and_factset():
    fnd = {
        "0700-HK": {
            "fy1_eps_mean": 20.0, "fy2_eps_mean": 24.0,
            "fy1_eps_mean_4wk_ago": 19.0, "fy1_eps_stddev": 1.0,
            "fy1_eps_num_est": 30, "factset_sector": "Communication Services",
            "factset_industry": "Interactive Media & Services",
        }
    }
    data = _populate(["0700-HK"], fundamentals=fnd)
    out = wing.parse_weekly_workbook(data, "data.xlsx")
    f = out["fundamentals"]["0700-HK"]
    assert f["fy1_eps_mean"] == 20.0 and f["fy2_eps_mean"] == 24.0
    assert f["fy1_eps_mean_4wk_ago"] == 19.0 and f["fy1_eps_stddev"] == 1.0
    assert f["factset_sector"] == "Communication Services"
    assert f["factset_industry"] == "Interactive Media & Services"
    assert out["meta"].get("n_fundamentals") == 1


def test_ingest_blank_and_na_fundamentals_become_none():
    fnd = {
        "9988-HK": {
            "fy1_eps_mean": "#N/A", "fy2_eps_mean": "",
            "fy1_eps_stddev": "NA", "factset_sector": "Consumer Discretionary",
        }
    }
    data = _populate(["9988-HK"], fundamentals=fnd)
    out = wing.parse_weekly_workbook(data, "data.xlsx")
    f = out["fundamentals"]["9988-HK"]
    assert f.get("fy1_eps_mean") is None
    assert f.get("fy2_eps_mean") is None
    assert f.get("fy1_eps_stddev") is None
    assert f["factset_sector"] == "Consumer Discretionary"


def test_two_column_only_template_still_parses_without_fundamentals():
    # Old price/volume-only template (no fundamentals block) must still parse.
    data = _populate(["0700-HK", "9988-HK"], include_fundamentals=False)
    out = wing.parse_weekly_workbook(data, "data.xlsx")
    assert set(out["tickers"].keys()) == {"0700-HK", "9988-HK"}
    # fundamentals present but empty for every ticker -> {} (graceful).
    assert out.get("fundamentals", {}) == {} or all(
        not v for v in out.get("fundamentals", {}).values()
    )


# ---------------------------------------------------------------------------
# Task 4 — valuation fwd P/E = price / FY1 EPS; EPS<=0/missing -> None
# ---------------------------------------------------------------------------
def test_valuation_fwd_pe_positive():
    v = M.valuation_metrics(360.0, {"fy1_eps_mean": 20.0})
    assert math.isclose(v["fwd_pe"], 18.0, rel_tol=1e-9)
    assert v["fy1_eps_mean"] == 20.0


def test_valuation_fwd_pe_none_when_eps_nonpositive_or_missing():
    assert M.valuation_metrics(360.0, {"fy1_eps_mean": 0.0})["fwd_pe"] is None
    assert M.valuation_metrics(360.0, {"fy1_eps_mean": -5.0})["fwd_pe"] is None
    assert M.valuation_metrics(360.0, {})["fwd_pe"] is None
    assert M.valuation_metrics(None, {"fy1_eps_mean": 20.0})["fwd_pe"] is None


# ---------------------------------------------------------------------------
# Task 4 — EPS momentum: revision abs/pct/dir, zero-base guard, dispersion
# ---------------------------------------------------------------------------
def test_earnings_momentum_revision_up():
    m = M.earnings_momentum({
        "fy1_eps_mean": 21.0, "fy1_eps_mean_4wk_ago": 20.0,
        "fy1_eps_stddev": 2.1, "fy1_eps_num_est": 25,
    })
    assert math.isclose(m["revision_abs"], 1.0, rel_tol=1e-9)
    assert math.isclose(m["revision_pct"], 0.05, rel_tol=1e-9)
    assert m["revision_dir"] == "up"
    assert math.isclose(m["dispersion"], 2.1 / 21.0, rel_tol=1e-9)
    assert m["num_est"] == 25


def test_earnings_momentum_revision_down_and_flat():
    down = M.earnings_momentum({"fy1_eps_mean": 18.0, "fy1_eps_mean_4wk_ago": 20.0})
    assert down["revision_dir"] == "down" and down["revision_abs"] < 0
    flat = M.earnings_momentum({"fy1_eps_mean": 20.0, "fy1_eps_mean_4wk_ago": 20.0})
    assert flat["revision_dir"] == "flat"


def test_earnings_momentum_zero_base_guard_and_missing():
    # Zero prior base -> revision_pct None (no div-by-zero), dir still computable.
    z = M.earnings_momentum({"fy1_eps_mean": 1.0, "fy1_eps_mean_4wk_ago": 0.0})
    assert z["revision_pct"] is None
    assert z["revision_dir"] == "up"
    # Missing prior -> all revision fields None.
    miss = M.earnings_momentum({"fy1_eps_mean": 20.0})
    assert miss["revision_abs"] is None and miss["revision_dir"] is None
    # Dispersion needs both stddev and non-zero fy1.
    assert M.earnings_momentum({"fy1_eps_stddev": 1.0})["dispersion"] is None


# ---------------------------------------------------------------------------
# Task 4 — attribution: stock-specific / sector-driven / solo + leave-one-out
# ---------------------------------------------------------------------------
def _row(sym, ret, sub=None, sec=None):
    return {"symbol": sym, "ret_1w": ret, "sub_industry": sub, "sector": sec}


def test_attribution_sector_driven_small_residual():
    # A tight cluster all up ~+6%; the name's residual is tiny -> Sector-driven.
    rows = [
        _row("A", 0.060, sub="Banks", sec="Financials"),
        _row("B", 0.058, sub="Banks", sec="Financials"),
        _row("C", 0.062, sub="Banks", sec="Financials"),
        _row("D", 0.061, sub="Banks", sec="Financials"),
    ]
    out = A.attribute_movers(rows, ["A"])
    a = out["A"]
    assert a["attribution"] == "Sector-driven"
    assert a["peer_group_used"] == "sub_industry"
    assert a["peer_count"] == 3  # leave-one-out excludes A itself
    # peer median of B,C,D = 0.061; residual = 0.060-0.061 = -0.001
    assert math.isclose(a["peer_median_1w"], 0.061, rel_tol=1e-9)
    assert math.isclose(a["residual_1w"], 0.060 - 0.061, abs_tol=1e-9)


def test_attribution_stock_specific_large_residual():
    # Peers quiet, the name rockets -> large residual -> Stock-specific.
    rows = [
        _row("A", 0.30, sub="Banks", sec="Financials"),
        _row("B", 0.005, sub="Banks", sec="Financials"),
        _row("C", 0.004, sub="Banks", sec="Financials"),
        _row("D", 0.006, sub="Banks", sec="Financials"),
    ]
    out = A.attribute_movers(rows, ["A"])
    a = out["A"]
    assert a["attribution"] == "Stock-specific"
    assert a["peer_count"] == 3
    assert a["residual_1w"] > 0.20


def test_attribution_solo_when_no_peers():
    rows = [_row("A", 0.10, sub="Solo Sub", sec="Solo Sec")]
    out = A.attribute_movers(rows, ["A"])
    a = out["A"]
    assert a["attribution"] == "Stock-specific"
    assert a["peer_group_used"] == A.SOLO
    assert a["peer_count"] == 0
    assert a["peer_median_1w"] is None and a["residual_1w"] is None


def test_attribution_rolls_up_to_sector_when_subindustry_thin():
    # Sub-industries unique (no sub peers) but all share a sector -> sector roll-up.
    rows = [
        _row("A", 0.06, sub="S1", sec="Energy"),
        _row("B", 0.05, sub="S2", sec="Energy"),
        _row("C", 0.07, sub="S3", sec="Energy"),
        _row("D", 0.06, sub="S4", sec="Energy"),
    ]
    out = A.attribute_movers(rows, ["A"])
    assert out["A"]["peer_group_used"] == "sector"
    assert out["A"]["peer_count"] == 3


# ---------------------------------------------------------------------------
# Task 5 — note renders fundamentals/attribution table when present; omits else
# ---------------------------------------------------------------------------
def _note_metrics(with_fund):
    per = {
        "0700-HK": {
            "ret_1w": 0.08, "rel_1w": 0.06, "vol_ratio": 1.4,
            "max_spike_ratio": 2.1, "sector": "Communication Services",
        }
    }
    if with_fund:
        per["0700-HK"].update({
            "fwd_pe": 18.0, "has_fundamentals": True,
            "momentum": {"revision_dir": "up", "revision_pct": 0.05,
                         "dispersion": 0.05},
            "attribution": {"attribution": "Stock-specific",
                            "peer_median_1w": 0.01, "residual_1w": 0.07,
                            "peer_group_used": "sub_industry", "peer_count": 4},
        })
    return {
        "asof": "2026-06-26",
        "movers": {"gainers_1w": [], "losers_1w": [], "vol_shift": [],
                   "vola_shift": [], "rel_leaders": [], "rel_laggards": []},
        "opportunities": {}, "hsi": {"loaded": True, "ret_1w": 0.01},
        "catalyst_names": ["0700-HK"], "per_ticker": per,
    }


def test_note_renders_fundamentals_attribution_when_present():
    tbl = N.render_fundamentals_table(_note_metrics(with_fund=True))
    assert "Fundamentals & attribution" in tbl
    assert "Fwd P/E" in tbl and "18.00x" in tbl
    assert "Stock-specific" in tbl and "up +5.0%" in tbl
    # And the deterministic backbone includes it.
    assert "Fundamentals & attribution" in N.render_metric_tables(_note_metrics(True))


def test_note_omits_fundamentals_table_when_absent():
    assert N.render_fundamentals_table(_note_metrics(with_fund=False)) == ""
    assert "Fundamentals & attribution" not in N.render_metric_tables(_note_metrics(False))


def test_catalyst_prompt_steers_query_by_attribution_tag():
    prompt = N.build_catalyst_prompt(_note_metrics(with_fund=True))
    assert "STOCK-SPECIFIC" in prompt
    assert "fwd P/E 18.00x" in prompt
    assert "FY1 EPS revised up" in prompt
    # No-fundamentals case keeps the prompt clean (no Attribution/Fundamentals).
    plain = N.build_catalyst_prompt(_note_metrics(with_fund=False))
    assert "Attribution:" not in plain and "Fundamentals:" not in plain
