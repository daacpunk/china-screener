"""Phase D weekly template generator: correct FDS formulas (price/volume/HSI),
fixed depth, valid multi-sheet workbook, batch split + ZIP."""
import io
import zipfile

from openpyxl import load_workbook

from app.weekly import template_gen as wtpl
from app.weekly import HSI_FACTSET_ID


def _load(b):
    return load_workbook(io.BytesIO(b))


def test_formulas_reference_cell_and_correct_depth():
    tf = wtpl.ticker_formulas("A2", wtpl.DEPTH)
    assert tf["close"] == f'=FDS(A2,"P_PRICE(0,-{wtpl.DEPTH}D,D)")'
    assert tf["volume"] == f'=FDS(A2,"P_VOLUME_DAY(0,-{wtpl.DEPTH}D,D)")'
    assert tf["date"].startswith('=FDS(A2,"JULIAN(P_PRICE(0,-') and ".dates)" in tf["date"]


def test_hsi_formulas_use_literal_id():
    hf = wtpl.hsi_formulas(wtpl.DEPTH)
    assert hf["close"] == f'=FDS("{HSI_FACTSET_ID}","P_PRICE(0,-{wtpl.DEPTH}D,D)")'
    assert HSI_FACTSET_ID == "180458"


def test_depth_floor_enforced():
    # An undersized requested depth is clamped up to the safe floor.
    b = wtpl.build_weekly_template(["0700-HK"], depth=10)
    wb = _load(b)
    ws = wb["0700-HK"]
    assert f"-{wtpl.DEPTH}D" in str(ws["C2"].value)


def test_build_has_instructions_hsi_manifest_and_ticker_sheets():
    tickers = ["0700-HK", "9988-HK", "0005-HK"]
    wb = _load(wtpl.build_weekly_template(tickers))
    names = set(wb.sheetnames)
    assert {"Instructions", "HSI", "Manifest"} <= names
    for t in tickers:
        assert t in names
    # HSI sheet: A2 literal id, C2 = HSI close formula
    hsi = wb["HSI"]
    assert str(hsi["A2"].value) == HSI_FACTSET_ID
    assert HSI_FACTSET_ID in str(hsi["C2"].value)
    # ticker sheet: A2 = literal ticker, C2 = price formula referencing A2
    ws = wb["0700-HK"]
    assert ws["A2"].value == "0700-HK"
    assert ws["C2"].value == f'=FDS(A2,"P_PRICE(0,-{wtpl.DEPTH}D,D)")'
    assert ws["D2"].value == f'=FDS(A2,"P_VOLUME_DAY(0,-{wtpl.DEPTH}D,D)")'


def test_instructions_contains_activate_macro():
    wb = _load(wtpl.build_weekly_template(["0700-HK"]))
    text = "\n".join(str(c.value) for row in wb["Instructions"].iter_rows() for c in row if c.value)
    assert "ActivateSpills" in text
    assert "Formula2" in text  # spill activation
    # Macro must skip the non-data sheets.
    assert "Instructions" in text and "Manifest" in text


def test_duplicate_ticker_sheet_names_disambiguated():
    # Two tickers that collapse to the same safe sheet name must not collide.
    wb = _load(wtpl.build_weekly_template(["A" * 40, "A" * 40 + "X"]))
    # all sheets unique (case-insensitive)
    lowered = [s.lower() for s in wb.sheetnames]
    assert len(lowered) == len(set(lowered))


def test_batch_split_and_zip():
    tickers = [f"T{i:03d}-HK" for i in range(170)]
    files = wtpl.build_weekly_templates_batched(tickers, batch_size=75)
    assert len(files) == 3  # 75 + 75 + 20
    # Each batch is a standalone workbook carrying its own HSI sheet.
    for fname, data in files:
        assert fname.endswith(".xlsx")
        wb = _load(data)
        assert "HSI" in wb.sheetnames
    zb = wtpl.zip_templates(files)
    zf = zipfile.ZipFile(io.BytesIO(zb))
    assert len(zf.namelist()) == 3


def test_total_ticker_sheets_across_batches():
    tickers = [f"T{i:03d}-HK" for i in range(170)]
    files = wtpl.build_weekly_templates_batched(tickers, batch_size=75)
    total = 0
    for _, data in files:
        wb = _load(data)
        total += len([s for s in wb.sheetnames if s not in ("Instructions", "HSI", "Manifest")])
    assert total == 170


def test_all_in_one_includes_every_ticker_above_batch_size():
    # all_in_one must NOT cap/split: > BATCH_SIZE tickers -> ONE workbook with
    # every ticker sheet + the single HSI + Instructions + Manifest sheets.
    n = wtpl.BATCH_SIZE + 30  # comfortably above the batch threshold
    tickers = [f"T{i:03d}-HK" for i in range(n)]
    wb = _load(wtpl.build_weekly_template(tickers))
    names = set(wb.sheetnames)
    assert {"Instructions", "HSI", "Manifest"} <= names
    ticker_sheets = [s for s in wb.sheetnames if s not in ("Instructions", "HSI", "Manifest")]
    assert len(ticker_sheets) == n
    for t in tickers:
        assert t in names
    # Exactly one HSI sheet in the single file.
    assert sum(1 for s in wb.sheetnames if s == "HSI") == 1


def test_split_zip_each_file_has_hsi_and_instructions():
    # Split mode -> ZIP of multiple workbooks; EACH must independently carry the
    # HSI sheet + Instructions so it can be refreshed/uploaded standalone.
    tickers = [f"T{i:03d}-HK" for i in range(160)]
    files = wtpl.build_weekly_templates_batched(tickers, batch_size=50)
    zb = wtpl.zip_templates(files)
    zf = zipfile.ZipFile(io.BytesIO(zb))
    members = zf.namelist()
    assert len(members) >= 2  # multiple workbooks
    for member in members:
        wb = _load(zf.read(member))
        names = set(wb.sheetnames)
        assert "HSI" in names, f"{member} missing HSI sheet"
        assert "Instructions" in names, f"{member} missing Instructions"
        # HSI uses the literal benchmark id.
        assert str(wb["HSI"]["A2"].value) == HSI_FACTSET_ID
