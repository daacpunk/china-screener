"""Phase D multi-file merge: merge_weekly_parsed reconciliation (disjoint +
duplicate tickers, HSI dedupe, common as-of, error aggregation) plus a
route-level two-file upload that merges into one active snapshot."""
import io

import numpy as np
import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.weekly import ingest as wing
from app.weekly import snapshot_store as wsnap
from app.weekly import template_gen as wtpl
from app.weekly import HSI_FACTSET_ID

client = TestClient(app)


def _julian(dates):
    origin = pd.Timestamp("1899-12-30")
    return [(pd.Timestamp(d) - origin).days for d in dates]


def _hsi_recs(last_date="2026-06-26", n=130):
    dates = pd.bdate_range(end=last_date, periods=n)
    closes = list(np.linspace(20000, 21000, n))
    return [{"date": pd.Timestamp(d).date().isoformat(), "close": float(c)}
            for d, c in zip(dates, closes)]


def _tkr_recs(last_date="2026-06-26", n=130):
    dates = pd.bdate_range(end=last_date, periods=n)
    closes = list(np.linspace(100, 120, n))
    return [{"date": pd.Timestamp(d).date().isoformat(), "close": float(c),
             "volume": 1_000_000.0} for d, c in zip(dates, closes)]


def _parsed(tickers_map, hsi, asof, source="f.xlsx", error=None, partial=None):
    out = {
        "asof": asof, "n_stale": None, "stale": False,
        "tickers": tickers_map, "hsi": hsi,
        "partial": partial or [],
        "meta": {"source": source, "n_tickers": len(tickers_map)},
    }
    if error:
        out["error"] = error
    return out


def test_merge_disjoint_tickers_identical_hsi():
    hsi = _hsi_recs()
    a = _parsed({"0700-HK": _tkr_recs()}, hsi, "2026-06-26", source="a.xlsx")
    b = _parsed({"9988-HK": _tkr_recs()}, list(hsi), "2026-06-26", source="b.xlsx")
    m = wing.merge_weekly_parsed([a, b])
    assert set(m["tickers"].keys()) == {"0700-HK", "9988-HK"}
    # HSI deduped to a single series (not doubled).
    assert len(m["hsi"]) == len(hsi)
    dates = [r["date"] for r in m["hsi"]]
    assert dates == sorted(dates)
    assert len(dates) == len(set(dates))
    assert m["asof"] == "2026-06-26"
    assert m["meta"]["n_files"] == 2
    assert set(m["sources"]) == {"a.xlsx", "b.xlsx"}
    assert not m.get("error")


def test_merge_common_asof_uses_min_last_date():
    hsi = _hsi_recs(last_date="2026-06-26")
    a = _parsed({"0700-HK": _tkr_recs(last_date="2026-06-26")}, hsi, "2026-06-26")
    # second file's ticker only reaches an earlier date -> common date is earlier
    b = _parsed({"9988-HK": _tkr_recs(last_date="2026-06-19")}, list(hsi), "2026-06-19")
    m = wing.merge_weekly_parsed([a, b])
    assert m["asof"] == "2026-06-19"


def test_merge_duplicate_ticker_keeps_richer_series():
    hsi = _hsi_recs()
    rich = _tkr_recs(n=130)
    poor = _tkr_recs(n=10)
    a = _parsed({"0700-HK": poor}, hsi, "2026-06-26")
    b = _parsed({"0700-HK": rich}, list(hsi), "2026-06-26")
    m = wing.merge_weekly_parsed([a, b])
    assert len(m["tickers"]["0700-HK"]) == 130
    # order-independent: richer wins even if it comes first
    m2 = wing.merge_weekly_parsed([b, a])
    assert len(m2["tickers"]["0700-HK"]) == 130


def test_merge_skips_empty_error_dict_gracefully():
    hsi = _hsi_recs()
    good = _parsed({"0700-HK": _tkr_recs()}, hsi, "2026-06-26", source="good.xlsx")
    bad = _parsed({}, [], None, source="bad.xlsx",
                  error="Could not read the workbook: boom")
    m = wing.merge_weekly_parsed([good, bad])
    assert "0700-HK" in m["tickers"]
    assert m["hsi"]
    assert m["asof"] == "2026-06-26"
    # per-file error surfaced as a soft, non-blocking warning
    assert m.get("error") and "boom" in m["error"]


def test_merge_all_empty_surfaces_error():
    a = _parsed({}, [], None, source="a.xlsx", error="bad a")
    b = _parsed({}, [], None, source="b.xlsx", error="bad b")
    m = wing.merge_weekly_parsed([a, b])
    assert not m["tickers"] and not m["hsi"]
    assert m.get("error")


def test_merge_single_element_is_identity_like():
    hsi = _hsi_recs()
    a = _parsed({"0700-HK": _tkr_recs()}, hsi, "2026-06-26", source="solo.xlsx")
    m = wing.merge_weekly_parsed([a])
    assert set(m["tickers"].keys()) == {"0700-HK"}
    assert len(m["hsi"]) == len(hsi)
    assert m["asof"] == "2026-06-26"
    assert m["meta"]["n_files"] == 1


def test_merge_hsi_dedupe_prefers_non_null_close():
    a = _parsed({}, [{"date": "2026-06-26", "close": None}], None)
    b = _parsed({}, [{"date": "2026-06-26", "close": 20500.0}], None)
    m = wing.merge_weekly_parsed([a, b])
    assert len(m["hsi"]) == 1
    assert m["hsi"][0]["close"] == 20500.0


# ----------------------------- route-level -----------------------------------
def _populate_one(tickers, last_date="2026-06-26", n_bars=130):
    """Build the real template and fill it FactSet-style -> .xlsx bytes."""
    b = wtpl.build_weekly_template(tickers)
    wb = load_workbook(io.BytesIO(b))
    dates = pd.bdate_range(end=last_date, periods=n_bars)
    jul = _julian(dates)
    for name in wb.sheetnames:
        if name in ("Instructions", "Manifest"):
            continue
        ws = wb[name]
        if name == "HSI":
            closes = list(np.linspace(20000, 21000, n_bars))
            for i, (jd, cl) in enumerate(zip(jul, closes)):
                ws.cell(row=2 + i, column=2, value=jd)
                ws.cell(row=2 + i, column=3, value=float(cl))
        else:
            closes = list(np.linspace(100, 120, n_bars))
            for i, (jd, cl) in enumerate(zip(jul, closes)):
                ws.cell(row=2 + i, column=2, value=jd)
                ws.cell(row=2 + i, column=3, value=float(cl))
                ws.cell(row=2 + i, column=4, value=1_000_000.0)
    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()


def test_route_two_file_upload_merges_one_snapshot(temp_db):
    f1 = _populate_one(["0700-HK"])
    f2 = _populate_one(["9988-HK"])
    r = client.post(
        "/weekly/data/upload",
        files=[
            ("files", ("batch1.xlsx", f1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("batch2.xlsx", f2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
        follow_redirects=False,
    )
    assert r.status_code == 303
    snap = wsnap.get_active()
    assert snap is not None
    data = snap["data"]
    assert set(data["tickers"].keys()) == {"0700-HK", "9988-HK"}
    assert data["hsi"]
    # HSI deduped to a single series despite appearing in both files.
    hd = [x["date"] for x in data["hsi"]]
    assert len(hd) == len(set(hd))
    assert data["meta"]["n_files"] == 2


def test_route_single_file_still_works(temp_db):
    f1 = _populate_one(["0700-HK"])
    r = client.post(
        "/weekly/data/upload",
        files=[("files", ("solo.xlsx", f1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        follow_redirects=False,
    )
    assert r.status_code == 303
    snap = wsnap.get_active()
    assert set(snap["data"]["tickers"].keys()) == {"0700-HK"}


def test_route_bad_file_redirects_with_error(temp_db):
    r = client.post(
        "/weekly/data/upload",
        files=[("files", ("notes.txt", b"not a workbook", "text/plain"))],
        follow_redirects=False,
    )
    assert r.status_code == 303
    assert "err=" in r.headers.get("location", "")
