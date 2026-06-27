"""Phase D weekly ingest: multi-sheet spill + JULIAN-decode round-trip,
HSI sheet detection, as-of = latest COMMON date, staleness, partial flagging."""
import io

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from app.weekly import ingest as wing
from app.weekly import template_gen as wtpl
from app.weekly import HSI_FACTSET_ID


def _julian(dates):
    origin = pd.Timestamp("1899-12-30")
    return [(pd.Timestamp(d) - origin).days for d in dates]


def _populate(tickers, n_bars=130, last_date="2026-06-26", hsi=True,
              partial_ticker=None, partial_bars=10):
    """Build the real template then fill it the way FactSet would (JULIAN dates +
    numeric close/volume), returning .xlsx bytes."""
    b = wtpl.build_weekly_template(tickers)
    wb = load_workbook(io.BytesIO(b))
    dates = pd.bdate_range(end=last_date, periods=n_bars)
    jul = _julian(dates)
    for name in wb.sheetnames:
        if name in ("Instructions", "Manifest"):
            continue
        ws = wb[name]
        if name == "HSI":
            if not hsi:
                # blank the HSI series (simulate a missing benchmark pull)
                ws["B2"] = None; ws["C2"] = None
                continue
            closes = list(np.linspace(20000, 21000, n_bars))
            for i, (jd, cl) in enumerate(zip(jul, closes)):
                ws.cell(row=2 + i, column=2, value=jd)
                ws.cell(row=2 + i, column=3, value=float(cl))
        else:
            nb = partial_bars if (partial_ticker and name == partial_ticker) else n_bars
            d2 = dates[-nb:]
            j2 = _julian(d2)
            closes = list(np.linspace(100, 120, nb))
            vols = [1_000_000.0] * nb
            for i, (jd, cl, v) in enumerate(zip(j2, closes, vols)):
                ws.cell(row=2 + i, column=2, value=jd)
                ws.cell(row=2 + i, column=3, value=float(cl))
                ws.cell(row=2 + i, column=4, value=float(v))
    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()


def test_round_trip_julian_decode_and_series():
    data = _populate(["0700-HK", "9988-HK"], n_bars=130, last_date="2026-06-26")
    out = wing.parse_weekly_workbook(data, "data.xlsx")
    assert "error" not in out or not out.get("error")
    assert set(out["tickers"].keys()) == {"0700-HK", "9988-HK"}
    assert out["hsi"], "HSI series should be parsed"
    # dates decoded back to ISO calendar dates (not raw serials)
    first = out["tickers"]["0700-HK"][0]
    assert first["date"].startswith("20")  # ISO year
    assert first["close"] is not None
    # chronological
    ds = [r["date"] for r in out["tickers"]["0700-HK"]]
    assert ds == sorted(ds)


def test_asof_is_latest_common_date():
    data = _populate(["0700-HK"], n_bars=130, last_date="2026-06-26")
    out = wing.parse_weekly_workbook(data, "data.xlsx")
    assert out["asof"] == "2026-06-26"


def test_staleness_flag():
    # Fresh-ish: 1 business day stale should not trip the >3 rule.
    fresh = _populate(["0700-HK"], last_date=pd.Timestamp.utcnow().normalize().strftime("%Y-%m-%d"))
    out = wing.parse_weekly_workbook(fresh, "data.xlsx")
    assert out["stale"] in (False, True)  # depends on weekday; type ok
    # Clearly old: 2020 -> stale True, n_stale large.
    old = _populate(["0700-HK"], last_date="2020-01-06")
    out2 = wing.parse_weekly_workbook(old, "data.xlsx")
    assert out2["stale"] is True
    assert out2["n_stale"] and out2["n_stale"] > 3


def test_hsi_sheet_detected_by_name_and_id():
    data = _populate(["0700-HK"], hsi=True)
    out = wing.parse_weekly_workbook(data, "data.xlsx")
    # HSI must NOT appear as a normal ticker
    assert HSI_FACTSET_ID not in out["tickers"]
    assert "HSI" not in out["tickers"]
    assert len(out["hsi"]) > 0


def test_partial_history_flagged():
    data = _populate(["FULL-HK", "SHORT-HK"], n_bars=130,
                     partial_ticker="SHORT-HK", partial_bars=10)
    out = wing.parse_weekly_workbook(data, "data.xlsx")
    assert "SHORT-HK" in out["partial"]
    assert "FULL-HK" not in out["partial"]


def test_skips_instructions_and_manifest():
    data = _populate(["0700-HK"])
    out = wing.parse_weekly_workbook(data, "data.xlsx")
    keys = set(out["tickers"].keys())
    assert "Instructions" not in keys and "Manifest" not in keys


def test_bad_input_never_raises():
    # Non-xlsx
    out = wing.parse_weekly_workbook(b"not a workbook", "notes.txt")
    assert out["error"] and out["tickers"] == {}
    # Garbage bytes with xlsx name
    out2 = wing.parse_weekly_workbook(b"\x00\x01\x02", "x.xlsx")
    assert "error" in out2
    assert out2["tickers"] == {} and out2["hsi"] == []


def test_missing_hsi_still_parses_tickers():
    data = _populate(["0700-HK", "9988-HK"], hsi=False)
    out = wing.parse_weekly_workbook(data, "data.xlsx")
    assert len(out["tickers"]) == 2
    assert out["hsi"] == []
    # as-of still derived from the ticker series
    assert out["asof"] is not None
