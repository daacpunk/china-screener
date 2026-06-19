"""generate_formula + BOTH formula-layout generators + xlsx builder.

Corrected FQL: comma-separated date args INSIDE the field parentheses,
most-recent-first (0D,-250D,D); volume via P_VOLUME_DAY; Method B uses the
bullet-proof 0D-Nd single-date offset.
"""
import io
import json
from pathlib import Path

import openpyxl

from app import formula_gen as fg

DICT = {
    "formulas": {
        "price": {"fql_template": "P_PRICE({start},{end},{freq})"},
        "volume": {"fql_template": "P_VOLUME_DAY({start},{end},{freq})"},
    }
}


def test_generate_formula_basic():
    out = fg.generate_formula("9988-HK", "price", DICT, start="0D", end="-250D", freq="D")
    assert out == '=FDS("9988-HK", "P_PRICE(0D,-250D,D)")'
    # commas, not colons
    assert ":" not in out


def test_generate_formula_unknown_metric():
    out = fg.generate_formula("X", "nope", DICT)
    assert out.startswith("# ERROR: Unknown metric")


def test_generate_formula_partial_placeholders():
    # only some placeholders provided -> others left intact
    out = fg.generate_formula("X", "price", DICT, start="0D")
    assert "0D" in out and "{end}" in out


def test_method_a_timeseries():
    a = fg.method_a_timeseries_formulas("9988-HK", DICT, start="0D", end="-250D", freq="D")
    assert a["close"] == '=FDS("9988-HK", "P_PRICE(0D,-250D,D)")'
    assert a["volume"] == '=FDS("9988-HK", "P_VOLUME_DAY(0D,-250D,D)")'
    # comma form, NO colon
    assert "0D,-250D,D" in a["close"]
    assert "0D:-250D:D" not in a["close"]
    assert ":" not in a["close"] and ":" not in a["volume"]
    # volume uses P_VOLUME_DAY
    assert "P_VOLUME_DAY" in a["volume"]
    assert "P_DATE" in a["date"]


def test_method_a_defaults_rolling_from_today():
    # Defaults should be the rolling-from-today window.
    a = fg.method_a_timeseries_formulas("9988-HK", DICT)
    assert a["close"] == '=FDS("9988-HK", "P_PRICE(0D,-150D,D)")'
    assert a["volume"] == '=FDS("9988-HK", "P_VOLUME_DAY(0D,-150D,D)")'


def test_method_b_offset_grid():
    grid = fg.method_b_offset_grid(DICT, lookback=10, ticker_cell="$A$2", header_rows=3)
    assert len(grid) == 10
    first = grid[0]
    assert first["row"] == 4 and first["offset"] == 0
    # 0D-Nd offset form for price and volume
    assert first["relative_formula"] == '=FDS($A$2,"P_PRICE(0D-"&(ROW()-3)&"D)")'
    assert first["relative_volume_formula"] == '=FDS($A$2,"P_VOLUME_DAY(0D-"&(ROW()-3)&"D)")'
    assert "0D-" in first["relative_formula"]
    assert "P_VOLUME_DAY" in first["relative_volume_formula"]
    # explicit pattern references column B for the row
    assert first["explicit_date_formula"] == '=FDS($A$2,"P_PRICE("&B4&")")'


def test_build_workbook_method_a_opens():
    # Default: NO date column (efficient). close=col1, volume=col2.
    data = fg.build_formula_workbook(["9988-HK", "PDD-CN"], DICT, method="A",
                                     layout="per_ticker", lookback=150)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Instructions" in wb.sheetnames
    assert "9988-HK" in wb.sheetnames
    ws = wb["9988-HK"]
    assert [c.value for c in ws[1]] == ["close", "volume"]  # no date column
    close2 = str(ws.cell(row=2, column=1).value)
    close3 = str(ws.cell(row=3, column=1).value)
    volume2 = str(ws.cell(row=2, column=2).value)
    assert 'P_PRICE(0D)' in close2
    assert 'P_PRICE(0D-1D)' in close3
    assert "P_VOLUME_DAY" in volume2
    assert ws.max_row == 1 + 150
    assert ":" not in close2 and ":" not in close3


def test_build_workbook_method_a_with_date_column():
    # include_date=True restores the date column (col1) -> close col2, volume col3.
    data = fg.build_formula_workbook(["9988-HK"], DICT, method="A",
                                     layout="per_ticker", lookback=10, include_date=True)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["9988-HK"]
    assert [c.value for c in ws[1]] == ["date", "close", "volume"]
    assert 'P_DATE(0D)' in str(ws.cell(row=2, column=1).value)
    assert 'P_PRICE(0D)' in str(ws.cell(row=2, column=2).value)
    assert 'P_VOLUME_DAY(0D)' in str(ws.cell(row=2, column=3).value)


def test_method_a_grid_row_per_day():
    grid = fg.method_a_grid("9988-HK", DICT, lookback=5)
    assert len(grid) == 5
    # row 1 = today (0D), row 2 = 0D-1D, ...
    assert grid[0]["close_formula"] == '=FDS("9988-HK","P_PRICE(0D)")'
    assert grid[1]["close_formula"] == '=FDS("9988-HK","P_PRICE(0D-1D)")'
    assert grid[0]["volume_formula"] == '=FDS("9988-HK","P_VOLUME_DAY(0D)")'
    assert grid[4]["close_formula"] == '=FDS("9988-HK","P_PRICE(0D-4D)")'
    # no colon range form anywhere
    assert all(":" not in g["close_formula"] for g in grid)


def test_build_workbook_method_a_stacked():
    # Stacked = tidy LONG format: one row per (ticker, day) with explicit
    # single-date formulas (no array-spill). Tab-3 ingestible.
    data = fg.build_formula_workbook(["AAA", "BBB"], DICT, method="A",
                                     layout="stacked", lookback=10)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "AllTickers" in wb.sheetnames
    ws = wb["AllTickers"]
    # Default: no date column -> ticker, close, volume.
    assert [c.value for c in ws[1]] == ["ticker", "close", "volume"]
    # header + 2 tickers * 10 days = 21 rows
    assert ws.max_row == 1 + 2 * 10
    # first ticker block: row 2 = AAA today, explicit single-date formula (col2)
    assert ws.cell(row=2, column=1).value == "AAA"
    assert ws.cell(row=2, column=2).value == '=FDS("AAA","P_PRICE(0D)")'
    assert ws.cell(row=3, column=2).value == '=FDS("AAA","P_PRICE(0D-1D)")'
    # second ticker block starts at row 12
    assert ws.cell(row=12, column=1).value == "BBB"
    assert ws.cell(row=12, column=2).value == '=FDS("BBB","P_PRICE(0D)")'
    for r in range(2, ws.max_row + 1):
        assert ":" not in str(ws.cell(row=r, column=2).value)


def test_build_workbook_method_b_opens():
    data = fg.build_formula_workbook(["9988-HK"], DICT, method="B", lookback=20)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["9988-HK"]
    assert ws.cell(row=2, column=1).value == "9988-HK"  # A2 = ticker
    price = str(ws.cell(row=4, column=3).value)
    volume = str(ws.cell(row=4, column=4).value)
    # 0D-Nd offset form, P_VOLUME_DAY for volume
    assert "ROW()" in price and "0D-" in price
    assert "P_VOLUME_DAY" in volume and "0D-" in volume


# ---- bundled sample dictionary: corrected templates ----
def test_sample_dictionary_corrected_templates():
    p = Path(__file__).resolve().parent.parent / "sample_data" / "dictionary.json"
    data = json.loads(p.read_text())
    formulas = data["formulas"]
    assert data["version"] == "2.0.0"
    # price template: commas not colons
    price_tmpl = formulas["price"]["fql_template"]
    assert price_tmpl == "P_PRICE({start},{end},{freq})"
    assert ":" not in price_tmpl
    # volume uses P_VOLUME_DAY
    vol_tmpl = formulas["volume"]["fql_template"]
    assert "P_VOLUME_DAY" in vol_tmpl
    assert ":" not in vol_tmpl
    # no fake P_ADV_USD field in any fql_template (notes may mention it)
    all_templates = " ".join(v["fql_template"] for v in formulas.values())
    assert "P_ADV_USD" not in all_templates
    # corrected GICS field names
    assert formulas["sector"]["fql_template"].startswith("FG_GICS_SECTOR")
    assert "FG_GICS_SUB_IND" in formulas["sub_industry"]["fql_template"]


def test_sample_dictionary_generates_corrected_price():
    p = Path(__file__).resolve().parent.parent / "sample_data" / "dictionary.json"
    data = json.loads(p.read_text())
    out = fg.generate_formula("9988-HK", "price", data, start="0D", end="-250D", freq="D")
    assert out == '=FDS("9988-HK", "P_PRICE(0D,-250D,D)")'


# ---- ITEM 4: configurable price/volume metric keys ----
DICT_CUSTOM = {
    "formulas": {
        "px_last": {"fql_template": "PX_LAST({start},{end},{freq})"},
        "vol": {"fql_template": "VOL_TRADED({start},{end},{freq})"},
    }
}


def test_autodetect_metrics_custom_keys():
    auto = fg.autodetect_metrics(DICT_CUSTOM)
    assert auto["price_metric"] == "px_last"  # contains 'px'
    assert auto["volume_metric"] == "vol"     # contains 'vol'


def test_workbook_honors_custom_metric_keys():
    # Without configurable keys this fell back to generic P_PRICE/P_VOLUME_DAY.
    data = fg.build_formula_workbook(
        ["9988-HK"], DICT_CUSTOM, method="A", layout="per_ticker",
        price_metric="px_last", volume_metric="vol",
    )
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["9988-HK"]
    # default: no date column -> close col1, volume col2
    close = str(ws.cell(row=2, column=1).value)
    volume = str(ws.cell(row=2, column=2).value)
    # The user's fql_template root is used, NOT the generic fallback (grid form).
    assert "PX_LAST(0D)" in close
    assert "P_PRICE" not in close
    assert "VOL_TRADED(0D)" in volume
    assert "P_VOLUME" not in volume


def test_method_b_honors_custom_price_metric():
    grid = fg.method_b_offset_grid(DICT_CUSTOM, lookback=5, price_metric="px_last", volume_metric="vol")
    assert "PX_LAST" in grid[0]["relative_formula"]
    assert "P_PRICE" not in grid[0]["relative_formula"]
    assert "VOL_TRADED" in grid[0]["relative_volume_formula"]


def test_min_required_bars_defaults_and_floor():
    # With default params: max(60, 60+21+1=82, 26*3+9=87, 14*3.5=49, 20) = 87
    # *1.25 = ~109, above the 90 floor.
    d = fg.min_required_bars(None)
    assert 95 <= d <= 130
    # Tiny params still respect the hard floor.
    tiny = {"vol_window": 10, "horizon_b_start": 5, "macd_slow": 5, "macd_signal": 3,
            "rsi_length": 5, "min_bars": 20, "sma_length": 10}
    assert fg.min_required_bars(tiny) == 90
    # Bigger vol window pushes depth up.
    big = {"vol_window": 120, "horizon_b_start": 21, "macd_slow": 26, "macd_signal": 9,
           "rsi_length": 14, "min_bars": 60, "sma_length": 20}
    assert fg.min_required_bars(big) > fg.min_required_bars(None)


def test_workbook_has_no_empty_cached_value_elements():
    # Excel flags formula cells that carry an empty <v/> cached value (repair
    # prompt). Generated workbooks must not contain any.
    import zipfile
    for method, layout in [("A", "per_ticker"), ("A", "stacked"), ("B", "per_ticker")]:
        data = fg.build_formula_workbook(["9988-HK", "BD5CMC"], DICT, method=method,
                                         layout=layout, lookback=8)
        z = zipfile.ZipFile(io.BytesIO(data))
        for n in z.namelist():
            if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                xml = z.read(n).decode("utf-8")
                assert "</f><v/>" not in xml and "</f><v />" not in xml, (method, layout, n)
        # still reopenable + formula intact
        wb = openpyxl.load_workbook(io.BytesIO(data))
        sheet = "AllTickers" if layout == "stacked" else "9988-HK"
        assert any("FDS(" in str(c.value) for row in wb[sheet].iter_rows() for c in row)


# ---- SPILL layout: single spilling range formula per series ----
def _af(v):
    # ArrayFormula values expose .text; plain strings pass through.
    return getattr(v, "text", v)


def test_method_a_spill_formulas_cell_ref_and_julian():
    s = fg.method_a_spill_formulas(DICT, lookback=109, ticker_cell="A2")
    # References the ticker CELL (A2), NOT a literal string. Start is 0 (not 0D).
    assert s["close"] == '=FDS(A2,"P_PRICE(0,-109D,D)")'
    assert s["volume"] == '=FDS(A2,"P_VOLUME_DAY(0,-109D,D)")'
    # Date axis uses JULIAN(...dates) on the price series.
    assert s["date"] == '=FDS(A2,"JULIAN(P_PRICE(0,-109D,D).dates)")'
    assert "0,-109D,D" in s["close"] and ":" not in s["close"]
    assert '"P_PRICE(0D' not in s["close"]  # not the old 0D anchor
    assert "P_VOLUME_DAY" in s["volume"]


def test_method_a_spill_formulas_custom_dict_roots():
    s = fg.method_a_spill_formulas(DICT_CUSTOM, lookback=50,
                                   price_metric="px_last", volume_metric="vol")
    assert s["close"] == '=FDS(A2,"PX_LAST(0,-50D,D)")'
    assert "P_PRICE" not in s["close"]
    assert "VOL_TRADED(0,-50D,D)" in s["volume"]
    assert "JULIAN(PX_LAST(0,-50D,D).dates)" in s["date"]


def test_build_workbook_spill_layout_bcd_columns():
    data = fg.build_formula_workbook(["9988-HK", "PDD-CN"], DICT, method="A",
                                     layout="spill", lookback=109)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Instructions" in wb.sheetnames
    assert "9988-HK" in wb.sheetnames
    ws = wb["9988-HK"]
    assert [c.value for c in ws[1]] == ["ticker", "date", "close", "volume"]
    assert ws.cell(row=2, column=1).value == "9988-HK"  # A2 = ticker literal
    assert _af(ws.cell(row=2, column=2).value) == '=FDS(A2,"JULIAN(P_PRICE(0,-109D,D).dates)")'
    assert _af(ws.cell(row=2, column=3).value) == '=FDS(A2,"P_PRICE(0,-109D,D)")'
    assert _af(ws.cell(row=2, column=4).value) == '=FDS(A2,"P_VOLUME_DAY(0,-109D,D)")'
    assert ws.max_row == 2


def test_build_workbook_spill_is_default_layout():
    data = fg.build_formula_workbook(["9988-HK"], DICT, method="A", lookback=109)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["9988-HK"]
    assert ws.cell(row=2, column=1).value == "9988-HK"
    assert _af(ws.cell(row=2, column=3).value) == '=FDS(A2,"P_PRICE(0,-109D,D)")'
    assert ws.max_row == 2


def test_spill_formulas_written_as_dynamic_array():
    # Spill formulas must be native DYNAMIC arrays: plain <f> + cm="1" on the cell
    # + an xl/metadata.xml with fDynamic="1". This avoids both the legacy CSE
    # '{' (openpyxl ArrayFormula) and the implicit-intersection '@'.
    import zipfile
    data = fg.build_formula_workbook(["9988-HK"], DICT, method="A",
                                     layout="spill", lookback=109)
    z = zipfile.ZipFile(io.BytesIO(data))
    assert "xl/metadata.xml" in z.namelist()
    meta = z.read("xl/metadata.xml").decode("utf-8")
    assert 'fDynamic="1"' in meta
    assert "sheetMetadata" in z.read("[Content_Types].xml").decode("utf-8")
    assert "metadata.xml" in z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    found = False
    for n in z.namelist():
        if n.startswith("xl/worksheets/") and n.endswith(".xml"):
            xml = z.read(n).decode("utf-8")
            if "P_PRICE(0,-109D,D)" in xml and "<f>" in xml:
                found = True
                # plain <f> (NOT t="array"), cell carries cm="1"
                assert 't="array"' not in xml
                assert 'cm="1"' in xml
                assert ">@FDS" not in xml and "&gt;@FDS" not in xml
                # no legacy CSE curly braces around the formula
                assert "<f>{" not in xml
    assert found



def test_build_workbook_per_ticker_still_row_per_day_grid():
    # per_ticker remains the explicit row-per-day grid (no spill) fallback.
    data = fg.build_formula_workbook(["9988-HK"], DICT, method="A",
                                     layout="per_ticker", lookback=20)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["9988-HK"]
    assert [c.value for c in ws[1]] == ["close", "volume"]
    assert "P_PRICE(0D)" in str(ws.cell(row=2, column=1).value)
    assert "P_PRICE(0D-1D)" in str(ws.cell(row=3, column=1).value)
    assert ws.max_row == 1 + 20  # full grid, NOT a single spill row


def test_build_formula_workbooks_batched_splits_200_at_75():
    tickers = [f"T{i:03d}" for i in range(200)]
    files = fg.build_formula_workbooks_batched(tickers, DICT, method="A",
                                               layout="spill", lookback=109,
                                               batch_size=75)
    assert len(files) == 3  # ceil(200/75) = 3
    names = [f for f, _ in files]
    assert names[0] == "factset_formulas_method_A_batch_01_of_03.xlsx"
    assert names[1] == "factset_formulas_method_A_batch_02_of_03.xlsx"
    assert names[2] == "factset_formulas_method_A_batch_03_of_03.xlsx"
    # batch sizes: 75, 75, 50
    sizes = []
    for fname, data in files:
        wb = openpyxl.load_workbook(io.BytesIO(data))
        # all sheets minus the Instructions sheet = ticker count in the batch
        sizes.append(len(wb.sheetnames) - 1)
    assert sizes == [75, 75, 50]
    # Instructions sheet notes the batch range
    wb1 = openpyxl.load_workbook(io.BytesIO(files[0][1]))
    instr_text = " ".join(
        str(c.value) for row in wb1["Instructions"].iter_rows() for c in row if c.value)
    assert "Batch 1 of 3" in instr_text
    assert "T000..T074" in instr_text


def test_batched_workbooks_have_no_empty_cached_values():
    import zipfile
    tickers = [f"T{i:03d}" for i in range(160)]
    files = fg.build_formula_workbooks_batched(tickers, DICT, method="A",
                                               layout="spill", lookback=20,
                                               batch_size=75)
    assert len(files) == 3
    for fname, data in files:
        z = zipfile.ZipFile(io.BytesIO(data))
        for n in z.namelist():
            if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                xml = z.read(n).decode("utf-8")
                assert "</f><v/>" not in xml and "</f><v />" not in xml, (fname, n)


def test_zip_workbooks_packs_all_entries():
    import zipfile
    tickers = [f"T{i:03d}" for i in range(200)]
    files = fg.build_formula_workbooks_batched(tickers, DICT, batch_size=75)
    zbytes = fg.zip_workbooks(files)
    z = zipfile.ZipFile(io.BytesIO(zbytes))
    assert len(z.namelist()) == 3
    assert all(n.endswith(".xlsx") for n in z.namelist())


def test_spill_workbook_has_no_empty_cached_values():
    import zipfile
    data = fg.build_formula_workbook(["9988-HK", "BD5CMC"], DICT, method="A",
                                     layout="spill", lookback=109, include_date=True)
    z = zipfile.ZipFile(io.BytesIO(data))
    for n in z.namelist():
        if n.startswith("xl/worksheets/") and n.endswith(".xml"):
            xml = z.read(n).decode("utf-8")
            assert "</f><v/>" not in xml and "</f><v />" not in xml, n
