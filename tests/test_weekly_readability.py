"""Weekly note readability overhaul (Phase D follow-up):

  * template emits FG_COMPANY_NAME + FNI_BUS_DESC_CO under the fundamentals
    toggle, and omits them when the toggle is off;
  * ingest parses company_name / business_desc, strips ■/control artifacts,
    coerces NA -> None, and still parses an old price/volume-only workbook;
  * metrics expose sector_median_fwd_pe, valuation_vs_sector (cheap / in line /
    rich) and ret_sigma on the mover entries;
  * note builds a KEY TAKEAWAYS box (AI via FakeProvider AND deterministic
    no-key), grouped plain-English movers with company names + sector tags +
    valuation-vs-sector anchors, a collapsed catalyst section (one "none found"
    line + a real-catalyst bullet + a low-confidence inferred group), and an
    inline glossary;
  * exporters render the new structure in md / html / docx / pdf with a clean
    glossary heading + definition string and no artifacts.
"""
import io
import zipfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from app import exporters
from app.llm.base import LLMProvider
from app.weekly import ingest as wing
from app.weekly import metrics as M
from app.weekly import note as N
from app.weekly import template_gen as wtpl


# ---------------------------------------------------------------------------
# Fake providers (mirror tests/test_weekly_note.py)
# ---------------------------------------------------------------------------
class FakeAnthropic(LLMProvider):
    name = "fake"  # not web-capable

    def __init__(self, *a, **k):
        super().__init__("fake-key", "fake-model")

    @property
    def available(self):
        return True

    def complete(self, prompt, **opts):
        # Deterministic, prompt-aware canned output so we can exercise the
        # takeaways box AND the catalyst collapse from a fake "model".
        if "KEY TAKEAWAYS" in prompt:
            return ("- AI takeaway one grounded in the figures.\n"
                    "- AI takeaway two.")
        if "OUTPUT FORMAT" in prompt:  # catalyst prompt
            return ("AAA-HK: no specific catalyst found\n"
                    "BBB-HK: likely moved with Energy on weak crude "
                    "(inferred, low confidence)")
        return "Canned section prose grounded in the provided figures."


class FakePerplexity(FakeAnthropic):
    name = "perplexity"  # web-capable


# ---------------------------------------------------------------------------
# Shared metrics fixture: two names with fundamentals (one cheap retailer with
# an EPS upgrade, one energy name with an EPS cut), plus a benchmark.
# ---------------------------------------------------------------------------
def _metrics_with_fundamentals():
    dates = pd.bdate_range("2026-01-01", periods=130).strftime("%Y-%m-%d").tolist()

    def recs(p, v):
        return [{"date": d, "close": float(x), "volume": float(y)}
                for d, x, y in zip(dates, p, v)]

    up = list(np.linspace(100, 118, 125)) + [119, 121, 123, 126, 140]
    dn = list(np.linspace(100, 90, 130))
    flat = list(np.linspace(100, 100, 130))
    # All three names share the BROAD sector so the sector median is meaningful.
    # fwd_pe is derived in-app as latest_close / fy1_eps_mean:
    #   AAA last close ~140, eps 14.0 -> ~10x (cheap vs median)
    #   BBB last close ~90,  eps 3.0  -> ~30x (rich  vs median)
    #   CCC last close ~100, eps 5.0  -> ~20x (the anchoring median)
    fund = {
        "AAA-HK": {
            "company_name": "Alpha Retail", "business_desc": "A retailer.",
            "factset_sector": "Consumer Discretionary", "factset_industry": "Retail",
            "fy1_eps_mean": 14.0, "fy1_eps_mean_4wk_ago": 12.0,
            "fy1_eps_num_est": 12,
        },
        "BBB-HK": {
            "company_name": "Beta Energy", "business_desc": "An oil company.",
            "factset_sector": "Consumer Discretionary", "factset_industry": "Retail",
            "fy1_eps_mean": 3.0, "fy1_eps_mean_4wk_ago": 3.6,
            "fy1_eps_num_est": 8,
        },
        "CCC-HK": {
            "company_name": "Gamma Goods", "business_desc": "A wholesaler.",
            "factset_sector": "Consumer Discretionary", "factset_industry": "Retail",
            "fy1_eps_mean": 5.0, "fy1_eps_mean_4wk_ago": 5.0,
            "fy1_eps_num_est": 10,
        },
    }
    snap = {
        "asof": "2026-06-26", "stale": False, "n_stale": 1, "partial": [],
        "tickers": {
            "AAA-HK": recs(up, [1e6] * 125 + [3e6, 1e6, 1e6, 1e6, 1e6]),
            "BBB-HK": recs(dn, [1e6] * 130),
            "CCC-HK": recs(flat, [1e6] * 130),
        },
        "hsi": [{"date": d, "close": float(c)}
                for d, c in zip(dates, np.linspace(20000, 21000, 130))],
        "fundamentals": fund,
    }
    return M.compute_weekly_metrics(snap)


# ===========================================================================
# Task 2 — template emits company name + business description formulas
# ===========================================================================
def test_template_emits_company_name_and_business_desc_under_toggle():
    ff = wtpl.fundamental_formulas("A2")
    assert ff["company_name"] == '=FDS(A2,"FG_COMPANY_NAME")'
    assert ff["business_desc"] == '=FDS(A2,"FNI_BUS_DESC_CO(ALL,1)")'

    wb = load_workbook(io.BytesIO(
        wtpl.build_weekly_template(["0700-HK"], include_fundamentals=True)))
    ws = wb["0700-HK"]
    cells = [str(ws.cell(row=r, column=wtpl.FUND_FORMULA_COL).value)
             for r in range(1, ws.max_row + 1)]
    blob = "\n".join(c for c in cells if c)
    assert "FG_COMPANY_NAME" in blob
    assert "FNI_BUS_DESC_CO(ALL,1)" in blob


def test_template_omits_fundamentals_when_toggle_off():
    wb = load_workbook(io.BytesIO(
        wtpl.build_weekly_template(["0700-HK"], include_fundamentals=False)))
    ws = wb["0700-HK"]
    blob = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "FG_COMPANY_NAME" not in blob
    assert "FNI_BUS_DESC_CO" not in blob


# ===========================================================================
# Task 3 — ingest parses names/desc, strips artifacts, NA->None, old files OK
# ===========================================================================
def test_ingest_cleans_text_artifacts_and_na():
    assert wing.clean_text_artifacts("Alpha\u25a0 Holdings\ufeff") == "Alpha Holdings"
    assert wing.clean_text_artifacts("\u25a0\u25a0\u25a0") is None
    # NA markers on a TEXT fundamental coerce to None.
    assert wing._clean_fund_value("@NA", "company_name") is None
    assert wing._clean_fund_value("  ", "business_desc") is None
    # A real text value survives (artifact stripped).
    assert wing._clean_fund_value("Tencent\u25a0", "company_name") == "Tencent"
    # Numeric fundamentals still coerce to float.
    assert wing._clean_fund_value("18.5", "fwd_pe") == 18.5


def _populate_with_fundamentals():
    """Build the real template and fill price/volume + the company-name /
    business-desc fundamental cells (with an embedded ■ artifact)."""
    tickers = ["0700-HK"]
    wb = load_workbook(io.BytesIO(wtpl.build_weekly_template(tickers)))
    dates = pd.bdate_range(end="2026-06-26", periods=130)
    origin = pd.Timestamp("1899-12-30")
    jul = [(pd.Timestamp(d) - origin).days for d in dates]
    closes = list(np.linspace(100, 120, 130))
    ws = wb["0700-HK"]
    for i, (jd, cl) in enumerate(zip(jul, closes)):
        ws.cell(row=2 + i, column=2, value=jd)
        ws.cell(row=2 + i, column=3, value=float(cl))
        ws.cell(row=2 + i, column=4, value=1_000_000.0)
    # Fill the fundamentals value column: match labels back to our two text rows.
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=wtpl.FUND_LABEL_COL).value
        if label == "Company name":
            ws.cell(row=r, column=wtpl.FUND_FORMULA_COL, value="Tencent\u25a0 Holdings")
        elif label == "Business description":
            ws.cell(row=r, column=wtpl.FUND_FORMULA_COL, value="Internet services.")
    # HSI series so it parses cleanly.
    hsi = wb["HSI"]
    for i, (jd, cl) in enumerate(zip(jul, np.linspace(20000, 21000, 130))):
        hsi.cell(row=2 + i, column=2, value=jd)
        hsi.cell(row=2 + i, column=3, value=float(cl))
    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()


def test_ingest_parses_company_name_and_business_desc():
    out = wing.parse_weekly_workbook(_populate_with_fundamentals(), "data.xlsx")
    f = out["fundamentals"]["0700-HK"]
    assert f["company_name"] == "Tencent Holdings"  # ■ stripped
    assert f["business_desc"] == "Internet services."


def test_ingest_old_price_only_workbook_still_parses():
    # A template built WITHOUT fundamentals must still ingest fine (no crash,
    # fundamentals dict empty / absent for the ticker).
    wb = load_workbook(io.BytesIO(
        wtpl.build_weekly_template(["0700-HK"], include_fundamentals=False)))
    dates = pd.bdate_range(end="2026-06-26", periods=130)
    origin = pd.Timestamp("1899-12-30")
    jul = [(pd.Timestamp(d) - origin).days for d in dates]
    ws = wb["0700-HK"]
    for i, (jd, cl) in enumerate(zip(jul, np.linspace(100, 120, 130))):
        ws.cell(row=2 + i, column=2, value=jd)
        ws.cell(row=2 + i, column=3, value=float(cl))
        ws.cell(row=2 + i, column=4, value=1_000_000.0)
    bio = io.BytesIO(); wb.save(bio)
    out = wing.parse_weekly_workbook(bio.getvalue(), "old.xlsx")
    assert "0700-HK" in out["tickers"]
    assert not (out.get("fundamentals") or {}).get("0700-HK")


# ===========================================================================
# Task 4 — metrics expose sector median / valuation-vs-sector / sigma
# ===========================================================================
def test_metrics_sector_median_and_valuation_vs_sector():
    # Direct unit checks on the pure helpers.
    rows = [{"sector": "Energy", "fwd_pe": 10.0},
            {"sector": "Energy", "fwd_pe": 20.0},
            {"sector": "Energy", "fwd_pe": 14.0}]
    med = M._sector_median_fwd_pe(rows)["Energy"]
    assert med == 14.0
    assert M._valuation_vs_sector(8.0, 14.0) == "cheap"     # < 0.85x
    assert M._valuation_vs_sector(14.0, 14.0) == "in line"
    assert M._valuation_vs_sector(20.0, 14.0) == "rich"     # > 1.15x
    # A single-name sector yields no reliable anchor.
    assert M._sector_median_fwd_pe([{"sector": "Lone", "fwd_pe": 9.0}])["Lone"] is None


def test_metrics_mover_entries_carry_anchors_and_sigma():
    m = _metrics_with_fundamentals()
    gainers = m["movers"]["gainers_1w"]
    assert gainers, "expected at least one gainer"
    g0 = gainers[0]
    # Names / sector carried onto the mover entry.
    assert g0.get("company_name")
    assert g0.get("sector") == "Consumer Discretionary"
    # Sector-median anchor populated (both names share the sector -> median).
    assert g0.get("sector_median_fwd_pe") is not None
    assert g0.get("valuation_vs_sector") in ("cheap", "in line", "rich")
    # Own-history sigma exposed.
    assert "ret_sigma" in g0
    # AAA (fwd P/E 8 vs median 10) should read cheap.
    per = m["per_ticker"]["AAA-HK"]
    assert per["valuation_vs_sector"] == "cheap"


# ===========================================================================
# Task 5 — note: takeaways box, grouped movers, catalyst collapse, glossary
# ===========================================================================
def test_no_key_note_has_deterministic_takeaways_movers_glossary():
    m = _metrics_with_fundamentals()
    md = N.generate_weekly_note(None, m, with_news=True)["markdown"]
    assert md.startswith("# Weekly Quant One-Pager")
    assert "## Key takeaways" in md
    assert "## What moved and why" in md
    # Grouped headings (plain English).
    assert "Gainers" in md
    # Company names + sector tags appear in the body, not bare codes only.
    assert "Alpha Retail" in md and "Consumer Discretionary" in md
    # Inline glossary with a definition string.
    assert "Glossary & methodology" in md
    assert "forward price-to-earnings" in md
    # No AI-only heading leaks into the no-key path.
    assert "## Data observations" not in md


def test_ai_note_has_takeaways_box_and_collapsed_catalysts():
    m = _metrics_with_fundamentals()
    note = N.generate_weekly_note(FakePerplexity(), m, with_news=True)
    md = note["markdown"]
    # KEY TAKEAWAYS box authored by the fake AI.
    assert "## Key takeaways" in md
    assert "AI takeaway one" in md
    # Catalysts collapsed: ONE "no catalyst" line + the low-confidence group.
    assert "## Catalysts (web)" in md
    assert "No company-specific news catalyst found for:" in md
    assert "Sector-driven (inferred, low confidence)" in md
    # No double-prefixed symbol in the inferred bullet.
    assert "BBB-HK: BBB-HK" not in md
    # Standard sections retained.
    assert "## Data observations" in md
    assert "## HSI macro view" in md
    assert "## Computed metrics" in md
    assert "Glossary & methodology" in md


def test_catalyst_collapse_groups_none_found_and_real_and_inferred():
    m = _metrics_with_fundamentals()
    raw = (
        "- AAA-HK: announced a major buyback program.\n"
        "- BBB-HK: no specific catalyst found\n"
    )
    out = N.collapse_catalysts(raw, m)
    # Real catalyst kept as its own bullet.
    assert "buyback program" in out
    # BBB collapsed into the single none-found line.
    assert "No company-specific news catalyst found for:" in out
    assert "BBB" in out


def test_grouped_movers_uses_valuation_anchor_and_plain_english():
    m = _metrics_with_fundamentals()
    body = N.render_grouped_movers(m)
    # Valuation phrasing surfaces "vs sector" for the cheap name.
    assert "vs sector" in body or "cheap" in body
    # Plain-English attribution language present.
    assert ("company-specific" in body) or ("sector" in body)


def test_takeaways_deterministic_never_empty():
    # Even on near-empty metrics the deterministic takeaways returns a bullet.
    out = N.render_takeaways_deterministic({})
    assert out.strip().startswith("-")


def test_hsi_prompt_does_not_restate_individual_movers():
    m = _metrics_with_fundamentals()
    p = N.build_hsi_prompt(m)
    assert "DO NOT restate" in p


# ===========================================================================
# Task 6 — exporters render the new structure cleanly; PDF valid; no artifacts
# ===========================================================================
def test_exporters_render_glossary_and_strip_artifacts():
    m = _metrics_with_fundamentals()
    # Inject an artifact into a rendered name to prove exporters strip it too.
    note = N.generate_weekly_note(FakePerplexity(), m, with_news=True)
    note["markdown"] = note["markdown"].replace("Alpha Retail", "Alpha\u25a0 Retail")

    md_bytes, _, _ = exporters.export(note, "md")
    assert b"\xe2\x96\xa0" not in md_bytes  # ■ utf-8 bytes gone
    md_txt = md_bytes.decode("utf-8")
    assert "Glossary & methodology" in md_txt
    assert "forward price-to-earnings" in md_txt

    html_bytes, _, _ = exporters.export(note, "html")
    html_txt = html_bytes.decode("utf-8")
    assert "\u25a0" not in html_txt
    assert "Glossary" in html_txt  # rendered as an <h3>

    docx_bytes, _, fname = exporters.export(note, "docx")
    assert fname.endswith(".docx") and len(docx_bytes) > 50
    # Inspect the docx XML for the glossary heading text + no artifact.
    with zipfile.ZipFile(io.BytesIO(docx_bytes)) as z:
        doc_xml = z.read("word/document.xml").decode("utf-8")
    assert "Glossary" in doc_xml
    assert "\u25a0" not in doc_xml

    pdf_bytes, _, _ = exporters.export(note, "pdf")
    assert pdf_bytes[:4] == b"%PDF"
    assert len(pdf_bytes) > 1000
