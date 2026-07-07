"""FactSet FQL formula generation.

Pure (no web/DB). Replicates the reference ``generate_formula`` exactly and
adds two layout builders plus a downloadable .xlsx builder.

Method A (preferred): per-ticker daily time-series block (date column + close
+ volume) using the corrected comma FQL, rolling from today, e.g.
=FDS("9988-HK","P_PRICE(0D,-250D,D)") and =FDS("9988-HK","P_VOLUME_DAY(0D,-250D,D)").

Method B (fallback): bullet-proof single-date offset grid using 0D-Nd, e.g.
=FDS($A$2,"P_PRICE(0D-"&(ROW()-3)&"D)") and P_VOLUME_DAY likewise.

Date args are COMMA-separated INSIDE the field parentheses (NOT a colon
"start:end:freq" string), most-recent-first (0D, then lookback). Volume uses
P_VOLUME_DAY (not P_VOLUME). ADV (USD) is computed in-app, not pulled.
"""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Reference function — replicated EXACTLY (generalised via kwargs).
# ---------------------------------------------------------------------------
def min_required_bars(params: Dict[str, Any] | None = None, buffer_pct: float = 0.25,
                      hard_floor: int = 90) -> int:
    """Compute the minimum CONTIGUOUS daily depth the screen actually needs.

    Indicators are consecutive-day calcs (Wilder RSI, MACD EMAs, rolling SMA,
    trailing vol window) so dates cannot be skipped — 'only needed dates' means
    the minimum contiguous lookback that warms every indicator. Binding terms:
      - trailing vol window reaching back past Horizon B's start
        (vol_window covers returns; Horizon B reaches day -21)
      - MACD warm-up for STABLE values (~3x slow EMA + signal)
      - RSI warm-up (~3.5x length)
      - the engine's hard min_bars floor
    A safety buffer is added on top, then a sensible hard floor.
    """
    p = dict(params or {})
    vol_window = int(p.get("vol_window", 60))
    h_b_start = int(p.get("horizon_b_start", 21))
    macd_slow = int(p.get("macd_slow", 26))
    macd_signal = int(p.get("macd_signal", 9))
    rsi_length = int(p.get("rsi_length", 14))
    min_bars = int(p.get("min_bars", 60))
    sma_length = int(p.get("sma_length", 20))

    # +1 because returns consume one bar; vol window must sit before latest moves.
    vol_need = vol_window + max(h_b_start, 5) + 1
    macd_need = macd_slow * 3 + macd_signal
    rsi_need = int(rsi_length * 3.5)
    base = max(min_bars, vol_need, macd_need, rsi_need, sma_length)
    depth = int(round(base * (1.0 + buffer_pct)))
    return max(hard_floor, depth)


def autodetect_metrics(dictionary: dict) -> Dict[str, str]:
    """Pick smart default price/volume metric keys from a dictionary.

    Price: first key whose name (case-insensitive) contains price/close/px.
    Volume: first key containing volume/vol. Falls back to the first key.
    """
    keys = list((dictionary or {}).get("formulas", {}).keys())
    first = keys[0] if keys else ""

    def _find(substrs: List[str]) -> str:
        for k in keys:
            kl = k.lower()
            if any(s in kl for s in substrs):
                return k
        return first

    return {
        "price_metric": _find(["price", "close", "px"]),
        "volume_metric": _find(["volume", "vol"]),
    }


def generate_formula(ticker: str, metric_key: str, dictionary: dict, **kwargs) -> str:
    formulas = dictionary["formulas"]
    if metric_key not in formulas:
        return f"# ERROR: Unknown metric '{metric_key}'"
    template = formulas[metric_key]["fql_template"]
    for key, value in kwargs.items():
        placeholder = "{" + key + "}"
        if placeholder in template:
            template = template.replace(placeholder, str(value))
    return f'=FDS("{ticker}", "{template}")'


# ---------------------------------------------------------------------------
# Method A — time-series block per ticker.
# ---------------------------------------------------------------------------
def method_a_timeseries_formulas(
    ticker: str,
    dictionary: dict,
    start: str = "0D",
    end: str = "-150D",
    freq: str = "D",
    price_metric: str = "price",
    volume_metric: str = "volume",
) -> Dict[str, str]:
    """Return the header FDS formulas for a daily date+close+volume block.

    Uses the corrected COMMA FQL form, most-recent-first (start=0D, end=-250D),
    e.g. P_PRICE(0D,-250D,D) and P_VOLUME_DAY(0D,-250D,D). FactSet returns the
    date axis automatically next to the price spill when a time-series formula
    is entered with a date range, so a single formula per series spills down.
    """
    formulas = dictionary.get("formulas", {})
    out: Dict[str, str] = {}
    if price_metric in formulas:
        out["close"] = generate_formula(
            ticker, price_metric, dictionary, start=start, end=end, freq=freq
        )
    else:
        out["close"] = f'=FDS("{ticker}", "P_PRICE({start},{end},{freq})")'
    if volume_metric in formulas:
        out["volume"] = generate_formula(
            ticker, volume_metric, dictionary, start=start, end=end, freq=freq
        )
    else:
        out["volume"] = f'=FDS("{ticker}", "P_VOLUME_DAY({start},{end},{freq})")'
    # date axis helper (best-effort; FactSet also spills the date column next to
    # the price spill automatically).
    out["date"] = f'=FDS("{ticker}", "P_DATE({start},{end},{freq})")'
    return out


def method_a_grid(
    ticker: str,
    dictionary: dict,
    lookback: int = 150,
    price_metric: str = "price",
    volume_metric: str = "volume",
    header_rows: int = 1,
    include_date: bool = False,
) -> List[Dict[str, Any]]:
    """Explicit row-per-day grid for one ticker (no array-spill dependency).

    Each trading day gets its own self-contained single-date FDS formula for
    date, close and volume using the 0D-Nd offset (today minus N trading days):
      row 1 (offset 0) = today (0D), row 2 = -1D, ... row N = -(N-1)D.
    This ALWAYS returns a full time series regardless of the FactSet add-in's
    dynamic-array support. Identifier is embedded directly in each formula so
    each sheet is self-contained.

    date  : =FDS("T","P_DATE(0D-N D)")        (best-effort; see note)
    close : =FDS("T","P_PRICE(0D-N D)")
    volume: =FDS("T","P_VOLUME_DAY(0D-N D)")
    """
    formulas = dictionary.get("formulas", {})
    price_fql = _fql_root(formulas, price_metric, "P_PRICE")
    vol_fql = _fql_root(formulas, volume_metric, "P_VOLUME_DAY")
    date_fql = _fql_root(formulas, "date_point", "P_DATE")
    rows: List[Dict[str, Any]] = []
    for i in range(lookback):
        off = f"0D-{i}D" if i else "0D"
        entry = {
            "row": header_rows + 1 + i,
            "offset": i,
            "close_formula": f'=FDS("{ticker}","{price_fql}({off})")',
            "volume_formula": f'=FDS("{ticker}","{vol_fql}({off})")',
        }
        # Date column is optional: dropping it removes ~1/3 of FDS calls. Dates
        # are reconstructed in-app from row order (row 1 = latest trading day).
        if include_date:
            entry["date_formula"] = f'=FDS("{ticker}","{date_fql}({off})")'
        rows.append(entry)
    return rows


# ---------------------------------------------------------------------------
# Spilling per-ticker layout — the fast default for Method A.
# ---------------------------------------------------------------------------
def _event_fql(formulas: dict) -> str:
    """FQL root for the next-event date metric (RTP_EARNINGS_RELEASE_DATE).

    Prefers a dictionary entry named like next_earnings / event / report; else
    falls back to RTP_EARNINGS_RELEASE_DATE (the live FactSet real-time earnings
    release date field, pulled via =FDSLIVE not =FDS).
    """
    for key in ("next_earnings", "event_date", "next_event", "earnings_date", "report_date"):
        if key in formulas:
            return _fql_root(formulas, key, "RTP_EARNINGS_RELEASE_DATE")
    # name-pattern scan
    for k in formulas:
        kl = k.lower()
        if any(s in kl for s in ("earn", "event", "report", "rep_dt")):
            return _fql_root(formulas, k, "RTP_EARNINGS_RELEASE_DATE")
    return "RTP_EARNINGS_RELEASE_DATE"


def _company_name_fql(formulas: dict) -> str:
    """FQL root for the company-name identifier (FG_COMPANY_NAME).

    Prefers a dictionary entry named like company_name / name / security; else
    falls back to FG_COMPANY_NAME (the authoritative point-in-time name field).
    """
    for key in ("company_name", "name", "security", "company"):
        if key in formulas:
            return _fql_root(formulas, key, "FG_COMPANY_NAME")
    for k in formulas:
        if "name" in k.lower() or "company" in k.lower():
            return _fql_root(formulas, k, "FG_COMPANY_NAME")
    return "FG_COMPANY_NAME"


def _fds_escape(template: str) -> str:
    """Escape a raw FQL template for embedding as the SECOND string arg of
    ``=FDS(<id>,"<template>")``.

    Inside a quoted =FDS field string, any nested double-quote must be DOUBLED
    (the same convention the add-in / the P_-with-string fields use). A template
    stored in the dictionary as ``FCA_EVENT_DATE(0,"CASH_DIST","EXDATE","YYYYMMDD")``
    therefore emits as ``FCA_EVENT_DATE(0,""CASH_DIST"",""EXDATE"",""YYYYMMDD"")``.
    Templates with no embedded quotes (e.g. ``FF_DIV_COM_CF(ANN,0,,,RF)``) are
    returned unchanged. (RTP_ earnings fields do not go through this =FDS escaper
    at all — they are emitted via ``fdslive_formula`` / ``=FDSLIVE``.)
    """
    return str(template).replace('"', '""')


def _dict_template(formulas: dict, key: str, default: str) -> str:
    """Full FQL template (with args) for a metric key, from the active dictionary;
    falls back to ``default`` when the key is absent. Unlike ``_fql_root`` this
    keeps the argument list intact (needed for the event-date fields)."""
    entry = (formulas or {}).get(key) or {}
    tmpl = entry.get("fql_template") if isinstance(entry, dict) else None
    return tmpl or default


def _earnings_template(formulas: dict) -> str:
    """Field name for the earnings-release DATE, preferring the dictionary entry.

    This is a live real-time (RTP_) field pulled via ``=FDSLIVE`` (NOT ``=FDS``):
    ``RTP_EARNINGS_RELEASE_DATE`` returns an int like ``20260831`` (YYYYMMDD).
    It has no date args and no nested quotes.
    """
    for key in ("next_earnings", "earnings_date_next", "earnings_date", "report_date"):
        if key in (formulas or {}):
            return _dict_template(formulas, key, "RTP_EARNINGS_RELEASE_DATE")
    return "RTP_EARNINGS_RELEASE_DATE"


def _earnings_status_template(formulas: dict) -> str:
    """Field name for the earnings-release STATUS, preferring the dictionary entry.

    Also a live real-time (RTP_) field pulled via ``=FDSLIVE``:
    ``RTP_EARNINGS_RELEASE_STATUS`` returns text like ``"Projected"`` /
    ``"Confirmed"``. No date args, no nested quotes.
    """
    for key in ("earnings_release_status", "earnings_status", "release_status"):
        if key in (formulas or {}):
            return _dict_template(formulas, key, "RTP_EARNINGS_RELEASE_STATUS")
    return "RTP_EARNINGS_RELEASE_STATUS"


def _ex_div_template(formulas: dict) -> str:
    """Full ex-dividend template, preferring the dictionary entry."""
    for key in ("ex_dividend_date", "ex_div_date", "exdate"):
        if key in (formulas or {}):
            return _dict_template(
                formulas, key, 'FCA_EVENT_DATE(0,"CASH_DIST","EXDATE","YYYYMMDD")'
            )
    return 'FCA_EVENT_DATE(0,"CASH_DIST","EXDATE","YYYYMMDD")'


def _event_date_formula(ticker_or_cell: str, template: str) -> str:
    """Single-cell =FDS(...,"<template>") for an event-date field.

    ``ticker_or_cell`` may be a literal ticker (embedded as "9988-HK") or a bare
    cell reference (e.g. A2). The nested ``template`` is escaped with doubled
    double-quotes so any embedded string args round-trip inside the =FDS field.
    """
    tc = str(ticker_or_cell)
    esc = _fds_escape(template)
    if re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", tc):
        return f'=FDS({tc},"{esc}")'
    return f'=FDS("{tc}","{esc}")'


def fdslive_formula(ticker_or_cell: str, field: str) -> str:
    """Single-cell ``=FDSLIVE(<id>,"<FIELD>")`` for a live real-time (RTP_) field.

    RTP fields (e.g. RTP_EARNINGS_RELEASE_DATE / RTP_EARNINGS_RELEASE_STATUS) are
    pulled with the live ``=FDSLIVE`` function, NOT the standard ``=FDS``. They
    take no date args and carry no nested quotes, so the field name is wrapped in
    simple double-quotes only (no doubling).

    ``ticker_or_cell`` follows the SAME literal-ticker vs cell-ref convention as
    the =FDS helpers: a bare cell reference (A2, $A$2) is used unquoted
    (``=FDSLIVE(A2,"...")``); anything else is a literal identifier wrapped in
    double quotes (``=FDSLIVE("9988-HK","...")``).
    """
    tc = str(ticker_or_cell)
    fld = str(field)
    if re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", tc):
        return f'=FDSLIVE({tc},"{fld}")'
    return f'=FDSLIVE("{tc}","{fld}")'


def earnings_date_formula(ticker_or_cell: str, dictionary: dict) -> str:
    """Single-cell earnings-release DATE via the LIVE ``=FDSLIVE`` function using
    the dictionary field name (default ``RTP_EARNINGS_RELEASE_DATE``), e.g.
    ``=FDSLIVE(A2,"RTP_EARNINGS_RELEASE_DATE")`` -> ``20260831`` (YYYYMMDD int)."""
    formulas = (dictionary or {}).get("formulas", {})
    return fdslive_formula(ticker_or_cell, _earnings_template(formulas))


def earnings_status_formula(ticker_or_cell: str, dictionary: dict) -> str:
    """Single-cell earnings-release STATUS via the LIVE ``=FDSLIVE`` function using
    the dictionary field name (default ``RTP_EARNINGS_RELEASE_STATUS``), e.g.
    ``=FDSLIVE(A2,"RTP_EARNINGS_RELEASE_STATUS")`` -> text like ``"Projected"``."""
    formulas = (dictionary or {}).get("formulas", {})
    return fdslive_formula(ticker_or_cell, _earnings_status_template(formulas))


def ex_dividend_formula(ticker_or_cell: str, dictionary: dict) -> str:
    """Single-cell ex-dividend =FDS(...) using the dictionary template, emitting
    the doubled-quote form, e.g.
    ``=FDS("9988-HK","FCA_EVENT_DATE(0,""CASH_DIST"",""EXDATE"",""YYYYMMDD"")")``."""
    formulas = (dictionary or {}).get("formulas", {})
    return _event_date_formula(ticker_or_cell, _ex_div_template(formulas))


def company_name_formula(ticker_or_cell: str, dictionary: dict) -> str:
    """Single-cell =FDS(...,"FG_COMPANY_NAME") for one ticker.

    ``ticker_or_cell`` may be a literal ticker (embedded as "9988-HK") or a cell
    reference (e.g. A2). FG_COMPANY_NAME takes no date arg, so this is a single
    point-in-time value, matching the main screen's double-quote FDS escaping,
    e.g. =FDS("9988-HK","FG_COMPANY_NAME") or =FDS(A2,"FG_COMPANY_NAME").
    """
    formulas = (dictionary or {}).get("formulas", {})
    fql = _company_name_fql(formulas)
    tc = str(ticker_or_cell)
    # A bare cell reference (e.g. A2, $A$2) is used unquoted; anything else is a
    # literal identifier and gets wrapped in double quotes.
    if re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", tc):
        return f'=FDS({tc},"{fql}")'
    return f'=FDS("{tc}","{fql}")'


def method_a_spill_formulas(
    dictionary: dict,
    lookback: int = 109,
    price_metric: str = "price",
    volume_metric: str = "volume",
    ticker_cell: str = "A2",
    include_event: bool = False,
    include_name: bool = False,
    include_events: bool = False,
) -> Dict[str, str]:
    """Return ONE spilling range FDS formula per series, referencing a ticker CELL.

    The FactSet add-in spills a single date-range formula down its column, so we
    emit ~2-3 calls/ticker (date + price + volume) instead of one call per cell.

    Verified against the user's live FactSet add-in:
      - Reference the ticker CELL (A2), NOT an embedded literal string:
            =FDS(A2,"P_PRICE(0,-109D,D)")
      - Start anchor is ``0`` (NOT ``0D``): ``(0,-<N>D,D)``.
      - The date axis uses JULIAN on the price series' .dates accessor:
            =FDS(A2,"JULIAN(P_PRICE(0,-109D,D).dates)")
      - These MUST be written as dynamic-array (spilling) formulas so Excel does
        not prepend the implicit-intersection ``@`` (which would force a single
        value and kill the spill). See _write_spill_formula / array markup.

    Column convention on each ticker's own sheet: B=date, C=close, D=volume,
    with the ticker literal in A2.

    FQL roots come from the active dictionary's templates (via _fql_root) so
    custom dictionaries work.

    Returns a dict with keys 'date', 'close', 'volume' (formulas as strings).
    """
    formulas = dictionary.get("formulas", {})
    price_fql = _fql_root(formulas, price_metric, "P_PRICE")
    vol_fql = _fql_root(formulas, volume_metric, "P_VOLUME_DAY")
    rng = f"0,-{lookback}D,D"
    price_expr = f"{price_fql}({rng})"
    out = {
        "date": f'=FDS({ticker_cell},"JULIAN({price_expr}.dates)")',
        "close": f'=FDS({ticker_cell},"{price_expr}")',
        "volume": f'=FDS({ticker_cell},"{vol_fql}({rng})")',
    }
    if include_event:
        ev_fql = _event_fql(formulas)
        # Single-value next-event date (NOT a spill).
        out["event"] = f'=FDS({ticker_cell},"{ev_fql}(0)")'
    if include_events:
        # Verified single-value event fields (NOT spills). Earnings date + status
        # are LIVE RTP_ fields via =FDSLIVE; ex-dividend is an =FDS FCA_EVENT_DATE
        # pull with doubled-quote escaping (unchanged).
        out["earnings_date"] = earnings_date_formula(ticker_cell, dictionary)
        out["earnings_status"] = earnings_status_formula(ticker_cell, dictionary)
        out["ex_dividend_date"] = ex_dividend_formula(ticker_cell, dictionary)
    if include_name:
        # Single-value company name (NOT a spill). No date arg.
        out["name"] = company_name_formula(ticker_cell, dictionary)
    return out


# ---------------------------------------------------------------------------
# Method B — offset grid fallback.
# ---------------------------------------------------------------------------
def _fql_root(formulas: dict, metric: str, default: str) -> str:
    """Return the FQL function root (text before the first '(') for a metric."""
    if metric in formulas:
        tmpl = formulas[metric].get("fql_template", default)
        return tmpl.split("(")[0] if "(" in tmpl else tmpl
    return default


def method_b_offset_grid(
    dictionary: dict,
    lookback: int = 150,
    price_metric: str = "price",
    volume_metric: str = "volume",
    ticker_cell: str = "$A$2",
    header_rows: int = 3,
) -> List[Dict[str, Any]]:
    """Return a list of rows describing the bullet-proof offset-grid layout.

    Each row is a self-contained single-date formula using the 0D-Nd offset
    (today minus N trading days), so row 1 = today and rows go further back —
    a rolling window.

    Each row: {row, offset, relative_formula, relative_volume_formula,
               explicit_date_formula}.
    relative price : =FDS($A$2,"P_PRICE(0D-"&(ROW()-3)&"D)")
    relative volume: =FDS($A$2,"P_VOLUME_DAY(0D-"&(ROW()-3)&"D)")
    explicit pattern: =FDS($A$2,"P_PRICE("&B2&")")  (date in column B)
    """
    formulas = dictionary.get("formulas", {})
    base_fql = _fql_root(formulas, price_metric, "P_PRICE")
    vol_fql = _fql_root(formulas, volume_metric, "P_VOLUME_DAY")
    rows: List[Dict[str, Any]] = []
    for i in range(lookback):
        sheet_row = header_rows + 1 + i
        rel = f'=FDS({ticker_cell},"{base_fql}(0D-"&(ROW()-{header_rows})&"D)")'
        rel_vol = f'=FDS({ticker_cell},"{vol_fql}(0D-"&(ROW()-{header_rows})&"D)")'
        exp = f'=FDS({ticker_cell},"{base_fql}("&B{sheet_row}&")")'
        rows.append(
            {
                "row": sheet_row,
                "offset": i,
                "relative_formula": rel,
                "relative_volume_formula": rel_vol,
                "explicit_date_formula": exp,
            }
        )
    return rows


# ---------------------------------------------------------------------------
# Downloadable .xlsx builder.
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1f2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, ncols: int, row: int = 1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def build_formula_workbook(
    tickers: List[str],
    dictionary: dict,
    method: str = "A",
    lookback: int = 150,
    start: str = "0D",
    end: str = "-150D",
    freq: str = "D",
    layout: str = "spill",  # 'spill' (default) | 'per_ticker' | 'stacked'
    price_metric: str = "price",
    volume_metric: str = "volume",
    include_date: bool = False,
    include_event: bool = False,
    include_name: bool = True,
    include_events: bool = True,
    batch_note: str = "",
) -> bytes:
    """Build the formula workbook as xlsx bytes.

    Method A:
      - spill (DEFAULT): one sheet per ticker; A2 = ticker literal and a SINGLE
        spilling range formula per series (price/volume, + date if requested)
        in row 2. The add-in fills the whole column — ~2 calls/ticker.
      - per_ticker: one sheet per ticker with an explicit row-per-day grid
        (no spill dependency; one self-contained =FDS per trading day).
      - stacked: a single Instructions sheet + one tidy-long sheet (all tickers).
    Method B:
      - one sheet per ticker with ticker cell + offset grid + date column.

    ``batch_note`` (when set) is added to the Instructions sheet, e.g.
    "Batch 1 of 8 — tickers AAA..ZZZ".
    """
    wb = Workbook()
    # Instructions sheet first
    info = wb.active
    info.title = "Instructions"
    instr = [
        ["FactSet Price-Series Formula Generator"],
        [""],
        [f"Method: {method}  |  Lookback: {lookback} td  |  Range: {start}..{end} freq {freq}"],
    ]
    if batch_note:
        instr += [[batch_note]]
    if include_name:
        instr += [
            [""],
            ['A "company_name" column is included: a single-value '
             '=FDS(...,"FG_COMPANY_NAME") per ticker (no date arg). Tab 3 reads it '
             'so the screen shows "Company Name (TICKER)". It is optional and does '
             'not affect the price/volume pull.'],
        ]
    if include_events:
        instr += [
            [""],
            ['THREE OPTIONAL event columns are included per ticker (single values, '
             'no spill): "earnings_date" + "earnings_status" (LIVE real-time RTP_ '
             'fields via =FDSLIVE) and "ex_dividend_date" (=FDS FCA_EVENT_DATE).'],
            ['  =FDSLIVE(A2,"RTP_EARNINGS_RELEASE_DATE")    -> 20260831 (YYYYMMDD int)'],
            ['  =FDSLIVE(A2,"RTP_EARNINGS_RELEASE_STATUS")  -> "Projected" / "Confirmed"'],
            ['Ex-dividend uses the standard =FDS function; ex-div returns YYYYMMDD '
             '(e.g. 20260526) and its nested strings use DOUBLED quotes:'],
            ['  =FDS(A2,"FCA_EVENT_DATE(0,""CASH_DIST"",""EXDATE"",""YYYYMMDD"")")'],
            ['Note: earnings uses =FDSLIVE (live), ex-div uses =FDS. Tab 3 decodes '
             'the dates to real dates, carries the earnings status text through, '
             'and flags a MECHANICAL_DISLOCATION when an ex-div / earnings date '
             'falls in the screen event window. They do NOT affect the '
             'price/volume pull and are backward-compatible.'],
        ]
    instr += [
        [""],
        ["HOW TO USE:"],
    ]
    if method.upper() == "A" and layout == "spill":
        instr += [
            ["SPILL LAYOUT — the formulas are stored as TEXT so Excel doesn't add"],
            ["the '@' implicit-intersection on open. Run the ONE-CLICK macro below"],
            ["once to activate every sheet's spills (it re-enters each formula via"],
            [".Formula2, exactly as if you typed it — so they spill, no '@')."],
            [""],
            ["STEP 1 — Activate the spills (one time):"],
            ["  a) Open this workbook in Excel with the FactSet add-in installed."],
            ["  b) Press Option+F11 (Mac) / Alt+F11 (Win) to open the VBA editor."],
            ["  c) Insert > Module, paste the macro below, press F5 (or Run)."],
            ["  d) Close the editor. Every ticker sheet now has LIVE spilling"],
            ["     formulas in B2:D2 that fill the whole column."],
            [""],
            ["STEP 2 — Let FactSet refresh, then upload the workbook in Tab 3"],
            ["  (Tab 3 forward-fills the A2 ticker down the spilled rows, and reads"],
            ["   JULIAN dates automatically)."],
            [""],
            ["---------- COPY THIS MACRO ----------"],
            ["Sub ActivateSpills()"],
            ["  Dim ws As Worksheet, c As Range, cols As Variant, i As Integer"],
            ["  cols = Array(2, 3, 4)   ' B=date, C=close, D=volume"],
            ["  For Each ws In ThisWorkbook.Worksheets"],
            ['    If ws.Name <> "Instructions" Then'],
            ["      For i = LBound(cols) To UBound(cols)"],
            ["        Set c = ws.Cells(2, cols(i))"],
            ['        If Left(c.Text, 1) = "=" Then'],
            ["          Dim f As String: f = c.Text"],
            ['          c.Value = ""            ' + "' clear the text"],
            ["          c.Formula2 = f          " + "' enter as dynamic array (spills)"],
            ["        End If"],
            ["      Next i"],
            ["    End If"],
            ["  Next ws"],
            ["End Sub"],
            ["---------- END MACRO ----------"],
            [""],
            [f"Formulas (N={lookback}, rolling from today, your add-in's syntax):"],
            [f'  B2: =FDS(A2,"JULIAN(P_PRICE(0,-{lookback}D,D).dates)")'],
            [f'  C2: =FDS(A2,"P_PRICE(0,-{lookback}D,D)")'],
            [f'  D2: =FDS(A2,"P_VOLUME_DAY(0,-{lookback}D,D)")'],
            ["~3 calls per ticker. Volume uses P_VOLUME_DAY. ADV is computed in-app."],
            [""],
            ["No macros? Manual fallback: click C2, press F2 then Enter to re-enter"],
            ["the formula (it will spill); repeat for B2/D2. Or use the 'Explicit"],
            ["row-per-day grid' layout, which needs no spill."],
        ]
    elif method.upper() == "A":
        instr += [
            ["1. Open this workbook in Excel with the FactSet add-in installed."],
            [f"2. Each ticker sheet has an EXPLICIT row-per-day grid: {lookback} rows,"],
            ["   one self-contained =FDS formula per trading day for date / close /"],
            ["   volume. This does NOT rely on array-spill, so a full time series"],
            ["   always returns (one value per row)."],
            ["3. Row 1 (offset 0D) = today / most-recent trading day; each row below"],
            ["   goes one trading day further back (0D-1D, 0D-2D, ...). Rolling: re-pull"],
            ["   anytime to refresh to the latest close."],
            ["4. Formulas: close = P_PRICE(0D-N D), volume = P_VOLUME_DAY(0D-N D),"],
            ["   date = P_DATE(0D-N D). Volume uses P_VOLUME_DAY (NOT P_VOLUME)."],
            ["5. 20D ADV (USD) is NOT pulled here — it is computed in-app (Tab 3) from"],
            ["   daily price * volume."],
            ["6. After refresh, export the resulting tidy data and upload in Tab 3."],
            ["   (Stacked layout = a single 'AllTickers' sheet in tidy long format:"],
            ["   columns ticker/date/close/volume, one row per ticker per day — also"],
            ["   explicit per-row formulas, no spill. Upload that sheet directly.)"],
            [""],
            ["Note: if P_DATE returns blank in your entitlement, the close/volume"],
            ["columns still work; dates can be reconstructed from the row offset (row 1"],
            ["= latest trading day) or pull dates via Method B's explicit-date column."],
            [""],
            ["Troubleshooting (no data): use commas not colons inside the field;"],
            ["P_VOLUME_DAY not P_VOLUME; check identifier format (e.g. 9988-HK, BD5CMC);"],
            ["use 0D-first (most-recent-first) order."],
        ]
    else:
        instr += [
            ["1. Open in Excel with the FactSet add-in installed."],
            ["2. Put the ticker in cell A2 of each sheet (already populated)."],
            ["3. Column C holds relative-offset price formulas using the 0D-Nd form,"],
            ["   e.g. P_PRICE(0D-N D); row 1 = today (0D), increasing rows go back."],
            ["4. Column D holds relative-offset volume formulas via P_VOLUME_DAY(0D-N D)."],
            ["5. Column E holds explicit-date price formulas referencing dates in column B;"],
            ["   fill column B with dates if using explicit mode, else use columns C/D."],
            ["6. Rolling window: 0D = today, looking back N trading days. Re-pull anytime"],
            ["   to refresh to the latest close. 20D ADV (USD) is computed in-app (Tab 3)."],
            ["7. Refresh, then export tidy data and upload in Tab 3."],
            [""],
            ["Troubleshooting (no data): use commas not colons inside the field;"],
            ["P_VOLUME_DAY not P_VOLUME; check identifier format (e.g. 9988-HK, BD5CMC);"],
            ["use 0D-first (most-recent-first) order."],
        ]
    for r in instr:
        info.append(r)
    info["A1"].font = Font(bold=True, size=14)

    if method.upper() == "A":
        if layout == "spill":
            # Spilling layout: each ticker on its OWN sheet (no collision).
            # A2 = ticker literal; the spill formulas REFERENCE A2 and sit in row
            # 2: B2=date (JULIAN ...dates), C2=price, D2=volume. Each is written as
            # Store the spill formulas as TEXT (leading apostrophe) so Excel does
            # NOT process them on open (which injects the implicit-intersection
            # '@' that kills the spill). The one-click 'Activate spills' macro
            # (see Instructions sheet) re-enters them via .Formula2, exactly as if
            # typed by hand -> they spill with no '@'.
            for t in tickers:
                safe = _safe_sheet_name(t)
                ws = wb.create_sheet(safe)
                header = ["ticker", "date", "close", "volume"]
                if include_event:
                    header.append("next_event")
                if include_events:
                    header += ["earnings_date", "earnings_status", "ex_dividend_date"]
                if include_name:
                    header.append("company_name")
                ws.append(header)
                _style_header(ws, len(header))
                spill = method_a_spill_formulas(
                    dictionary, lookback=lookback,
                    price_metric=price_metric, volume_metric=volume_metric,
                    ticker_cell="A2", include_event=include_event,
                    include_name=include_name, include_events=include_events)
                ws.cell(row=2, column=1, value=t)  # A2 = ticker literal
                # Force TEXT storage so '=' is inert until the macro activates it.
                # Company name / event dates are appended AFTER B/C/D so the
                # spill-activation macro (which targets columns 2/3/4) is
                # untouched.
                cells = [(2, "date"), (3, "close"), (4, "volume")]
                next_col = 5
                if include_event:
                    # single next-event date value (not a spill).
                    cells.append((next_col, "event"))
                    next_col += 1
                if include_events:
                    # single-value event cells (not spills): earnings date +
                    # status (=FDSLIVE), then ex-dividend (=FDS).
                    cells.append((next_col, "earnings_date"))
                    next_col += 1
                    cells.append((next_col, "earnings_status"))
                    next_col += 1
                    cells.append((next_col, "ex_dividend_date"))
                    next_col += 1
                if include_name:
                    # single company-name value (not a spill).
                    cells.append((next_col, "name"))
                for col, key in cells:
                    c = ws.cell(row=2, column=col, value=spill[key])
                    c.data_type = "s"  # string, not formula
                widths = [(1, 14), (2, 16), (3, 30), (4, 32)]
                extra_col = 5
                if include_event:
                    widths.append((extra_col, 30))
                    extra_col += 1
                if include_events:
                    widths.append((extra_col, 24))
                    extra_col += 1
                    widths.append((extra_col, 20))
                    extra_col += 1
                    widths.append((extra_col, 40))
                    extra_col += 1
                if include_name:
                    widths.append((extra_col, 34))
                for col, w in widths:
                    ws.column_dimensions[get_column_letter(col)].width = w
        elif layout == "stacked":
            # Tidy LONG format: one row per (ticker, trading day) with explicit
            # single-date formulas (no array-spill). This is exactly the shape
            # Tab 3 ingests, all tickers stacked in a single sheet.
            ws = wb.create_sheet("AllTickers")
            header = (["ticker", "date", "close", "volume"] if include_date
                      else ["ticker", "close", "volume"])
            if include_events:
                header += ["earnings_date", "earnings_status", "ex_dividend_date"]
            if include_name:
                header.append("company_name")
            ws.append(header)
            _style_header(ws, len(header))
            for t in tickers:
                grid = method_a_grid(t, dictionary, lookback=lookback,
                                     price_metric=price_metric,
                                     volume_metric=volume_metric,
                                     include_date=include_date)
                for i, g in enumerate(grid):
                    if include_date:
                        row_vals = [t, g["date_formula"], g["close_formula"], g["volume_formula"]]
                    else:
                        row_vals = [t, g["close_formula"], g["volume_formula"]]
                    if include_events:
                        # Emit the event fields only on the first row of each
                        # ticker block (single values; Tab 3 forward-fills):
                        # earnings date + status (=FDSLIVE), ex-dividend (=FDS).
                        row_vals.append(earnings_date_formula(t, dictionary) if i == 0 else None)
                        row_vals.append(earnings_status_formula(t, dictionary) if i == 0 else None)
                        row_vals.append(ex_dividend_formula(t, dictionary) if i == 0 else None)
                    if include_name:
                        # Emit the name FDS only on the first row of each ticker
                        # block (one point-in-time value; Tab 3 forward-fills).
                        row_vals.append(company_name_formula(t, dictionary) if i == 0 else None)
                    ws.append(row_vals)
            if include_date:
                widths = [(1, 14), (2, 24), (3, 30), (4, 32)]
                extra_col = 5
            else:
                widths = [(1, 14), (2, 30), (3, 32)]
                extra_col = 4
            if include_events:
                widths.append((extra_col, 24))
                extra_col += 1
                widths.append((extra_col, 20))
                extra_col += 1
                widths.append((extra_col, 40))
                extra_col += 1
            if include_name:
                widths.append((extra_col, 34))
            for col, w in widths:
                ws.column_dimensions[get_column_letter(col)].width = w
        else:
            for t in tickers:
                safe = _safe_sheet_name(t)
                ws = wb.create_sheet(safe)
                header = ["date", "close", "volume"] if include_date else ["close", "volume"]
                base_ncol = 3 if include_date else 2
                event_col = base_ncol + 1  # first extra column after price/vol block
                if include_events:
                    header += ["earnings_date", "earnings_status", "ex_dividend_date"]
                if include_name:
                    header.append("company_name")
                ws.append(header)
                _style_header(ws, len(header))
                # Explicit row-per-day grid: one self-contained FDS formula per
                # trading day. No reliance on array-spill, so a full ~N-day time
                # series always returns. Row 2 = today, row 3 = -1D, ...
                grid = method_a_grid(t, dictionary, lookback=lookback,
                                     price_metric=price_metric,
                                     volume_metric=volume_metric,
                                     include_date=include_date)
                for g in grid:
                    r = g["row"]
                    if include_date:
                        ws.cell(row=r, column=1, value=g["date_formula"])
                        ws.cell(row=r, column=2, value=g["close_formula"])
                        ws.cell(row=r, column=3, value=g["volume_formula"])
                    else:
                        ws.cell(row=r, column=1, value=g["close_formula"])
                        ws.cell(row=r, column=2, value=g["volume_formula"])
                extra_col = event_col
                if include_events:
                    # Single-value event cells in row 2 (this sheet's ticker):
                    # earnings date + status (=FDSLIVE), then ex-dividend (=FDS).
                    ws.cell(row=2, column=extra_col,
                            value=earnings_date_formula(t, dictionary))
                    ws.cell(row=2, column=extra_col + 1,
                            value=earnings_status_formula(t, dictionary))
                    ws.cell(row=2, column=extra_col + 2,
                            value=ex_dividend_formula(t, dictionary))
                    extra_col += 3
                if include_name:
                    # Single company-name value in row 2 (this sheet's ticker).
                    ws.cell(row=2, column=extra_col,
                            value=company_name_formula(t, dictionary))
                widths = [(1, 22), (2, 30), (3, 32)] if include_date else [(1, 30), (2, 32)]
                wcol = event_col
                if include_events:
                    widths.append((wcol, 24))
                    widths.append((wcol + 1, 20))
                    widths.append((wcol + 2, 40))
                    wcol += 3
                if include_name:
                    widths.append((wcol, 34))
                for col, w in widths:
                    ws.column_dimensions[get_column_letter(col)].width = w
    else:
        for t in tickers:
            safe = _safe_sheet_name(t)
            ws = wb.create_sheet(safe)
            hdr = ["ticker_cell", "date_col_B", "relative_price_C",
                   "relative_volume_D", "explicit_price_E"]
            if include_events:
                hdr += ["earnings_date", "earnings_status", "ex_dividend_date"]
            if include_name:
                hdr.append("company_name")
            ws.append(hdr)
            _style_header(ws, len(hdr))
            ws.cell(row=2, column=1, value=t)  # A2 = ticker
            grid = method_b_offset_grid(dictionary, lookback=lookback,
                                        price_metric=price_metric,
                                        volume_metric=volume_metric)
            for g in grid:
                r = g["row"]
                ws.cell(row=r, column=3, value=g["relative_formula"])
                ws.cell(row=r, column=4, value=g["relative_volume_formula"])
                ws.cell(row=r, column=5, value=g["explicit_date_formula"])
            # Extra columns start at F (col 6), after the offset-grid columns.
            extra_col = 6
            widths = [(1, 14), (2, 14), (3, 34), (4, 34), (5, 34)]
            if include_events:
                # Single-value event cells referencing the ticker cell $A$2:
                # earnings date + status (=FDSLIVE), then ex-dividend (=FDS).
                ws.cell(row=2, column=extra_col,
                        value=earnings_date_formula("$A$2", dictionary))
                ws.cell(row=2, column=extra_col + 1,
                        value=earnings_status_formula("$A$2", dictionary))
                ws.cell(row=2, column=extra_col + 2,
                        value=ex_dividend_formula("$A$2", dictionary))
                widths.append((extra_col, 24))
                widths.append((extra_col + 1, 20))
                widths.append((extra_col + 2, 40))
                extra_col += 3
            if include_name:
                # Single company-name value referencing the ticker cell.
                ws.cell(row=2, column=extra_col, value=company_name_formula("$A$2", dictionary))
                widths.append((extra_col, 34))
            for col, w in widths:
                ws.column_dimensions[get_column_letter(col)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    return _strip_empty_formula_values(bio.getvalue())


def _chunk(seq: List[Any], size: int) -> List[List[Any]]:
    size = max(1, int(size))
    return [seq[i:i + size] for i in range(0, len(seq), size)]


def build_formula_workbooks_batched(
    tickers: List[str],
    dictionary: dict,
    method: str = "A",
    lookback: int = 150,
    start: str = "0D",
    end: str = "-150D",
    freq: str = "D",
    layout: str = "spill",
    price_metric: str = "price",
    volume_metric: str = "volume",
    include_date: bool = False,
    include_event: bool = False,
    include_name: bool = True,
    include_events: bool = True,
    batch_size: int = 75,
) -> List[tuple]:
    """Split ``tickers`` into chunks of ``batch_size`` and build one workbook per
    chunk. Returns a list of ``(filename, xlsx_bytes)`` tuples.

    Each workbook is a full standalone file (its own Instructions sheet noting
    'Batch k of M — tickers X..Y'). ``_strip_empty_formula_values`` runs on every
    workbook (inside ``build_formula_workbook``) so none trigger Excel's repair
    prompt. Filenames: ``factset_formulas_method_A_batch_01_of_08.xlsx`` etc.
    """
    chunks = _chunk(list(tickers), batch_size)
    total = len(chunks)
    width = max(2, len(str(total)))
    out: List[tuple] = []
    for idx, chunk in enumerate(chunks, start=1):
        first, last = chunk[0], chunk[-1]
        note = f"Batch {idx} of {total} — tickers {first}..{last} ({len(chunk)} names)"
        data = build_formula_workbook(
            chunk, dictionary, method=method, lookback=lookback,
            start=start, end=end, freq=freq, layout=layout,
            price_metric=price_metric, volume_metric=volume_metric,
            include_date=include_date, include_event=include_event,
            include_name=include_name, include_events=include_events,
            batch_note=note,
        )
        fname = (f"factset_formulas_method_{method}_batch_"
                 f"{idx:0{width}d}_of_{total:0{width}d}.xlsx")
        out.append((fname, data))
    return out


def zip_workbooks(files: List[tuple]) -> bytes:
    """Assemble a list of ``(filename, bytes)`` into a single in-memory zip."""
    import zipfile

    bio = io.BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, data in files:
            zf.writestr(fname, data)
    return bio.getvalue()


def _strip_empty_formula_values(xlsx_bytes: bytes) -> bytes:
    """Remove empty cached-value elements (``<v/>``) that openpyxl writes after
    every formula cell.

    openpyxl emits ``<f>FORMULA</f><v/>`` for formula cells. An empty-but-present
    cached value on a formula that references an add-in function not yet loaded
    (``FDS``) makes Excel report 'We found a problem with some content...' and
    offer to repair on open. Stripping the empty ``<v/>`` yields a clean
    ``<f>FORMULA</f>`` cell that Excel opens without complaint and recalculates
    once the FactSet add-in is present. The XML stays well-formed.
    """
    import zipfile

    has_dynamic = False
    zin = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        names = set(zin.namelist())
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                txt = data.decode("utf-8")
                txt = txt.replace("</f><v/>", "</f>").replace("</f><v />", "</f>")
                # Mark bare =FDS(...) spill formulas as DYNAMIC arrays so Excel
                # spills them (no leading '{' CSE, no '@' implicit intersection).
                # Target only cells whose entire formula is a single FDS(...) call
                # (the spill layout) — NOT Method B's '&'-concatenated offset
                # formulas, which must stay plain.
                def _mark(m):
                    nonlocal has_dynamic
                    cell_open, formula = m.group(1), m.group(2)
                    if formula.startswith("FDS(") and "&" not in formula and "ROW(" not in formula:
                        has_dynamic = True
                        return f'{cell_open} cm="1"><f>{formula}</f>'
                    return m.group(0)
                txt = re.sub(r'(<c r="[A-Z]+\d+"[^>]*)><f>([^<]*)</f>', _mark, txt)
                data = txt.encode("utf-8")
            if item.filename == "[Content_Types].xml":
                t = data.decode("utf-8")
                if "sheetMetadata" not in t:
                    t = t.replace(
                        "</Types>",
                        '<Override PartName="/xl/metadata.xml" '
                        'ContentType="application/vnd.openxmlformats-officedocument.'
                        'spreadsheetml.sheetMetadata+xml"/></Types>',
                    )
                data = t.encode("utf-8")
            rels_name = "xl/_rels/workbook.xml.rels"
            if item.filename == rels_name and "metadata.xml" not in data.decode("utf-8"):
                t = data.decode("utf-8")
                t = t.replace(
                    "</Relationships>",
                    '<Relationship Id="rIdMeta1" '
                    'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                    'relationships/sheetMetadata" Target="metadata.xml"/></Relationships>',
                )
                data = t.encode("utf-8")
            zout.writestr(item, data)
        # Add the dynamic-array metadata part (only meaningful if we marked cells,
        # but harmless to always include when spill formulas are present).
        if has_dynamic and "xl/metadata.xml" not in names:
            zout.writestr("xl/metadata.xml", _DYNAMIC_ARRAY_METADATA)
    return out.getvalue()


_DYNAMIC_ARRAY_METADATA = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<metadata xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
    'xmlns:xlrd="http://schemas.microsoft.com/office/spreadsheetml/2017/richdata" '
    'xmlns:xda="http://schemas.microsoft.com/office/spreadsheetml/2017/dynamicarray">'
    '<metadataTypes count="1">'
    '<metadataType name="XLDAPR" minSupportedVersion="120000" copy="1" pasteAll="1" '
    'pasteValues="1" merge="1" splitFirst="1" rowColShift="1" clearFormats="1" '
    'clearComments="1" assign="1" coerce="1" cellMeta="1"/>'
    '</metadataTypes>'
    '<futureMetadata name="XLDAPR" count="1"><bk><extLst>'
    '<ext uri="{bdbb8cdc-fa1e-496e-a857-3c3f30c029c3}">'
    '<xda:dynamicArrayProperties fDynamic="1" fCollapsed="0"/></ext>'
    '</extLst></bk></futureMetadata>'
    '<cellMetadata count="1"><bk><rc t="1" v="0"/></bk></cellMetadata>'
    '</metadata>'
)


def _safe_sheet_name(name: str) -> str:
    bad = set('[]:*?/\\')
    s = "".join("_" if ch in bad else ch for ch in str(name))
    return s[:31] if s else "Sheet"
