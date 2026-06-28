"""Phase D weekly FactSet template generator.

Builds a downloadable .xlsx using the proven spill / text-formula + ActivateSpills
VBA approach (the exact pattern in app/formula_gen.py). One sheet per ticker:

    A2 = FactSet ticker literal
    B2 = =FDS(A2,"JULIAN(P_PRICE(0,-{depth}D,D).dates)")   (JULIAN dates)
    C2 = =FDS(A2,"P_PRICE(0,-{depth}D,D)")                  (close, spills)
    D2 = =FDS(A2,"P_VOLUME_DAY(0,-{depth}D,D)")             (volume, spills)

The HSI benchmark series is pulled ONCE on a dedicated "HSI" sheet, using the
FactSet identifier 180458 as a literal in the formula (NOT a cell reference, so
it survives even if the sheet is reordered):

    B2 = =FDS("180458","JULIAN(P_PRICE(0,-{depth}D,D).dates)")
    C2 = =FDS("180458","P_PRICE(0,-{depth}D,D)")

depth is a FIXED safe floor of 300 trading days. 300 td ~= 14 calendar months,
which always covers YTD (max ~252 td) plus the 60D context volatility window
(252 + 60 = 312 only at year-end; the YTD anchor is the most recent year-start,
which is at most ~252 td back, and 6M=126 + 60D=186, so 300 covers every
headline + context window with comfortable margin while keeping ~3 FDS calls
per ticker).

Formulas are stored as TEXT (leading data_type='s') so Excel does not inject the
implicit-intersection '@' on open; the ActivateSpills macro on the Instructions
sheet re-enters them via .Formula2 so they spill. A Manifest sheet lists every
column, its FDS formula, and the depth. Batched into a ZIP at >75 tickers.

Pure (no web / DB). Reuses formula_gen._strip_empty_formula_values / zip helper.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .. import formula_gen as fg
from . import HSI_FACTSET_ID

# Fixed safe floor: 300 trading days always covers YTD + 60D context window.
DEPTH = 300
BATCH_SIZE = 75

# Fundamentals block layout: single-cell point-in-time formulas placed in the
# columns F (label) / G (=FDS(...) as TEXT) so they never collide with the
# spilling B/C/D price/volume series. One row per field, starting at row 2.
FUND_LABEL_COL = 6   # F
FUND_FORMULA_COL = 7  # G


def _price_expr(depth: int) -> str:
    return f"P_PRICE(0,-{depth}D,D)"


def _vol_expr(depth: int) -> str:
    return f"P_VOLUME_DAY(0,-{depth}D,D)"


# ---------------------------------------------------------------------------
# Fundamentals — point-in-time single-cell FE_ESTIMATE FQL + FactSet sector formulas.
# These are the EXACT, user-tested Excel-FQL forms (see FE_ESTIMATE_SYNTAX).
# Each value is a single cell (NOT a spill). Forward P/E is NOT a native field —
# it is computed in-app downstream as latest_close / FY1 EPS mean.
# ---------------------------------------------------------------------------
# key -> (human label, FQL expression inside =FDS(A2,"..."))
FUNDAMENTAL_FIELDS: List[Tuple[str, str, str]] = [
    ("fy1_eps_mean", "FY1 EPS mean", "FE_ESTIMATE(EPS,MEAN,ANN_ROLL,+1,NOW,,,'')"),
    ("fy2_eps_mean", "FY2 EPS mean", "FE_ESTIMATE(EPS,MEAN,ANN_ROLL,+2,NOW,,,'')"),
    ("fy1_eps_mean_4wk_ago", "FY1 EPS mean (-20D, ~4wks ago)",
     "FE_ESTIMATE(EPS,MEAN,ANN_ROLL,+1,-20D,,,'')"),
    ("fy1_eps_stddev", "FY1 EPS stddev", "FE_ESTIMATE(EPS,STDDEV,ANN_ROLL,+1,NOW,,,'')"),
    ("fy1_eps_num_est", "FY1 EPS num_est", "FE_ESTIMATE(EPS,NEST,ANN_ROLL,+1,NOW,,,'')"),
    ("factset_sector", "FactSet sector", "FG_FACTSET_SECTOR"),
    ("factset_industry", "FactSet industry", "FG_FACTSET_IND"),
    # Point-in-time descriptors used to label each name in the note (company
    # name + a short business description). EXACT user-tested Excel-FQL forms.
    ("company_name", "Company name", "FG_COMPANY_NAME"),
    ("business_desc", "Business description", "FNI_BUS_DESC_CO(ALL,1)"),
]


# label text -> canonical key (used by ingest to map the populated label column).
FUNDAMENTAL_LABEL_TO_KEY: Dict[str, str] = {
    label: key for key, label, _expr in FUNDAMENTAL_FIELDS
}
FUNDAMENTAL_KEYS: List[str] = [key for key, _label, _expr in FUNDAMENTAL_FIELDS]
# string-valued (non-numeric) fundamental fields.
FUNDAMENTAL_TEXT_KEYS = {
    "factset_sector", "factset_industry", "company_name", "business_desc",
}


def fundamental_formulas(ticker_cell: str = "A2") -> Dict[str, str]:
    """Per-ticker point-in-time fundamental formulas referencing the ticker CELL.

    Returns {key: '=FDS(A2,"<FQL>")'} using the EXACT FE_ESTIMATE / FG_FACTSET
    forms (NEST statistic; FG_FACTSET_SECTOR/FG_FACTSET_IND — GICS pulls no data
    in this entitlement). These are single cells (point-in-time), not spills.
    """
    return {key: f'=FDS({ticker_cell},"{expr}")' for key, _label, expr in FUNDAMENTAL_FIELDS}


def ticker_formulas(ticker_cell: str = "A2", depth: int = DEPTH) -> Dict[str, str]:
    """Per-ticker spill formulas referencing the ticker CELL (A2)."""
    pe = _price_expr(depth)
    return {
        "date": f'=FDS({ticker_cell},"JULIAN({pe}.dates)")',
        "close": f'=FDS({ticker_cell},"{pe}")',
        "volume": f'=FDS({ticker_cell},"{_vol_expr(depth)}")',
    }


def hsi_formulas(depth: int = DEPTH) -> Dict[str, str]:
    """HSI benchmark spill formulas, using the 180458 identifier as a LITERAL."""
    pe = _price_expr(depth)
    return {
        "date": f'=FDS("{HSI_FACTSET_ID}","JULIAN({pe}.dates)")',
        "close": f'=FDS("{HSI_FACTSET_ID}","{pe}")',
    }


_INSTRUCTIONS_HEADER = "Weekly Quant One-Pager — FactSet Template (Phase D)"


def _instructions_rows(depth: int, n_tickers: int, batch_note: str = "",
                       include_fundamentals: bool = True) -> List[List[str]]:
    tf = ticker_formulas("A2", depth)
    hf = hsi_formulas(depth)
    fund_col = get_column_letter(FUND_FORMULA_COL)
    n_fund = len(FUNDAMENTAL_FIELDS)
    rows: List[List[str]] = [
        [_INSTRUCTIONS_HEADER],
        [""],
        [f"Depth: {depth} trading days (fixed floor — covers YTD + 60D vol window)."],
        [f"Tickers in this file: {n_tickers}.  HSI benchmark id: {HSI_FACTSET_ID}."],
    ]
    if batch_note:
        rows += [[batch_note]]
    rows += [
        [""],
        ["SPILL LAYOUT — formulas are stored as TEXT so Excel doesn't add the"],
        ["'@' implicit-intersection on open. Run the ONE-CLICK macro below once"],
        ["to activate every sheet's spills (it re-enters each formula via"],
        [".Formula2, exactly as if you typed it — so they spill, no '@')."],
        [""],
        ["STEP 1 — Activate the spills (one time):"],
        ["  a) Open this workbook in Excel with the FactSet add-in installed."],
        ["  b) Press Option+F11 (Mac) / Alt+F11 (Win) to open the VBA editor."],
        ["  c) Insert > Module, paste the macro below, press F5 (or Run)."],
        ["  d) Close the editor. Every sheet now has LIVE spilling formulas."],
        [""],
        ["STEP 2 — Let FactSet refresh, then upload the workbook in the Weekly"],
        ["  Note tab (it reads each ticker sheet + the HSI sheet, decodes the"],
        ["  JULIAN dates, and computes all metrics in-app)."],
        [""],
        ["---------- COPY THIS MACRO ----------"],
        ["Sub ActivateSpills()"],
        ["  Dim ws As Worksheet, c As Range, cols As Variant, i As Integer"],
        ["  Dim fr As Integer"],
        ["  cols = Array(2, 3, 4)   ' B=date, C=close, D=volume"],
        ["  For Each ws In ThisWorkbook.Worksheets"],
        ['    If ws.Name <> "Instructions" And ws.Name <> "Manifest" Then'],
        ["      For i = LBound(cols) To UBound(cols)"],
        ["        Set c = ws.Cells(2, cols(i))"],
        ['        If Left(c.Text, 1) = "=" Then'],
        ["          Dim f As String: f = c.Text"],
        ['          c.Value = ""            ' + "' clear the text"],
        ["          c.Formula2 = f          " + "' enter as dynamic array (spills)"],
        ["        End If"],
        ["      Next i"],
        [f"      ' fundamentals: single cells in column {fund_col} (rows 2..{1 + n_fund})"],
        [f"      For fr = 2 To {1 + n_fund}"],
        [f"        Set c = ws.Cells(fr, {FUND_FORMULA_COL})  ' column {fund_col}"],
        ['        If Left(c.Text, 1) = "=" Then'],
        ["          Dim ff As String: ff = c.Text"],
        ['          c.Value = """"'],
        ["          c.Formula2 = ff"],
        ["        End If"],
        ["      Next fr"],
        ["    End If"],
        ["  Next ws"],
        ["End Sub"],
        ["---------- END MACRO ----------"],
        [""],
        [f"Per-ticker formulas (N={depth}, A2 = FactSet ticker):"],
        [f"  B2: {tf['date']}"],
        [f"  C2: {tf['close']}"],
        [f"  D2: {tf['volume']}"],
        [""],
        ["HSI benchmark sheet (pulled once):"],
        [f"  B2: {hf['date']}"],
        [f"  C2: {hf['close']}"],
        [""],
        ["~3 calls per ticker + 2 for HSI. Volume uses P_VOLUME_DAY. All returns,"],
        ["volatility, momentum, and HSI-relative metrics are computed IN-APP."],
        [""],
        ["No macros? Manual fallback: click C2, press F2 then Enter to re-enter"],
        ["the formula (it will spill); repeat for B2/D2 on each sheet."],
    ]
    if include_fundamentals:
        ff = fundamental_formulas("A2")
        rows += [
            [""],
            [f"FUNDAMENTALS (point-in-time single cells, column {fund_col}, one per row):"],
            ["  Estimates use the FE_ESTIMATE FQL function (EPS consensus); sector/"],
            ["  industry classification via FG_FACTSET_SECTOR / FG_FACTSET_IND;"],
            ["  company name + business description via FG_COMPANY_NAME /"],
            ["  FNI_BUS_DESC_CO(ALL,1) (used to label each name in the note)."],
            ["  Availability depends on your FactSet entitlement — NA shows as"],
            ["  blank and is treated n/a downstream (the note falls back to the"],
            ["  bare ticker when a name is missing)."],
        ]
        for i, (key, label, _expr) in enumerate(FUNDAMENTAL_FIELDS):
            rows.append([f"  {fund_col}{2 + i} ({label}): {ff[key]}"])
        rows += [
            ["  Forward P/E is NOT a native field — computed in-app as"],
            ["  latest_close / FY1 EPS mean (n/a when EPS<=0 or missing). EPS"],
            ["  revision direction/magnitude is computed in-app from the current"],
            ["  FY1 mean vs the -20D (~4 weeks ago) FY1 mean."],
        ]
    return rows


def _manifest_rows(tickers: List[str], depth: int,
                   include_fundamentals: bool = True) -> List[List[str]]:
    tf = ticker_formulas("A2", depth)
    hf = hsi_formulas(depth)
    rows = [
        ["sheet", "column", "field", "fds_formula", "depth_td"],
        ["<each ticker>", "A", "ticker", "(literal in A2)", depth],
        ["<each ticker>", "B", "date (JULIAN)", tf["date"], depth],
        ["<each ticker>", "C", "close", tf["close"], depth],
        ["<each ticker>", "D", "volume", tf["volume"], depth],
        ["HSI", "A", "identifier", HSI_FACTSET_ID, depth],
        ["HSI", "B", "date (JULIAN)", hf["date"], depth],
        ["HSI", "C", "close", hf["close"], depth],
    ]
    if include_fundamentals:
        ff = fundamental_formulas("A2")
        cell = get_column_letter(FUND_FORMULA_COL)
        rows += [
            ["", "", "", "", ""],
            ["FUNDAMENTALS (point-in-time single cells; col G per ticker sheet).",
             "", "", "", ""],
            ["Availability of estimate / FactSet classification fields depends on "
             "your FactSet entitlement; NA shows as blank and is treated n/a downstream.",
             "", "", "", ""],
        ]
        for i, (key, label, _expr) in enumerate(FUNDAMENTAL_FIELDS):
            rows.append(["<each ticker>", f"{cell}{2 + i}",
                         f"{label} (point-in-time)", ff[key], ""])
        rows.append(["Company name / business description (FG_COMPANY_NAME / "
                     "FNI_BUS_DESC_CO) require the FactSet Fundamentals/Company "
                     "entitlement; NA shows as blank and the name falls back to "
                     "the bare ticker downstream.", "", "", "", ""])
        rows.append(["<computed in-app>", "", "forward P/E",
                     "latest_close / FY1 EPS mean (n/a if EPS<=0/missing)", ""])
    rows += [
        ["", "", "", "", ""],
        ["Tickers in this file:", "", "", "", len(tickers)],
    ]
    for t in tickers:
        rows.append([t, "", "", "", ""])
    return rows


def _write_fundamentals_block(ws) -> None:
    """Write the point-in-time fundamentals block (label in F, =FDS(...) TEXT in G)
    on a ticker sheet, one field per row starting at row 2. Stored as TEXT so
    Excel doesn't inject '@' on open; the ActivateSpills macro re-enters them."""
    ff = fundamental_formulas("A2")
    ws.cell(row=1, column=FUND_LABEL_COL, value="fundamental")
    ws.cell(row=1, column=FUND_FORMULA_COL, value="value (point-in-time)")
    for i, (key, label, _expr) in enumerate(FUNDAMENTAL_FIELDS):
        r = 2 + i
        ws.cell(row=r, column=FUND_LABEL_COL, value=label)
        c = ws.cell(row=r, column=FUND_FORMULA_COL, value=ff[key])
        c.data_type = "s"
    ws.column_dimensions[get_column_letter(FUND_LABEL_COL)].width = 30
    ws.column_dimensions[get_column_letter(FUND_FORMULA_COL)].width = 46


def build_weekly_template(
    tickers: List[str],
    depth: int = DEPTH,
    batch_note: str = "",
    include_fundamentals: bool = True,
) -> bytes:
    """Build a single weekly-template workbook (Instructions + Manifest + HSI +
    one sheet per ticker) as xlsx bytes. Formulas stored as TEXT; spill metadata
    added via the shared _strip_empty_formula_values pass.

    ``include_fundamentals`` (default True) adds a point-in-time fundamentals
    block (FE_ESTIMATE EPS consensus + GICS) to each ticker sheet. When False,
    the lean price/volume-only layout (the original Phase D behavior) is built.
    """
    depth = max(int(depth or DEPTH), DEPTH)
    wb = Workbook()

    # Instructions sheet (first).
    info = wb.active
    info.title = "Instructions"
    for r in _instructions_rows(depth, len(tickers), batch_note, include_fundamentals):
        info.append(r)
    info["A1"].font = Font(bold=True, size=14)

    # HSI benchmark sheet (pulled once, literal identifier).
    hsi = wb.create_sheet("HSI")
    hsi.append(["identifier", "date", "close"])
    fg._style_header(hsi, 3)
    hf = hsi_formulas(depth)
    hsi.cell(row=2, column=1, value=HSI_FACTSET_ID)
    for col, key in ((2, "date"), (3, "close")):
        c = hsi.cell(row=2, column=col, value=hf[key])
        c.data_type = "s"
    for col, w in ((1, 14), (2, 16), (3, 30)):
        hsi.column_dimensions[get_column_letter(col)].width = w

    # Per-ticker sheets.
    used_names = {"instructions", "hsi", "manifest"}
    for t in tickers:
        base = fg._safe_sheet_name(t)
        safe = base
        suffix = 1
        while safe.lower() in used_names:
            tail = f"_{suffix}"
            safe = base[: 31 - len(tail)] + tail
            suffix += 1
        used_names.add(safe.lower())
        ws = wb.create_sheet(safe)
        ws.append(["ticker", "date", "close", "volume"])
        fg._style_header(ws, 4)
        tf = ticker_formulas("A2", depth)
        ws.cell(row=2, column=1, value=t)  # A2 = ticker literal
        for col, key in ((2, "date"), (3, "close"), (4, "volume")):
            c = ws.cell(row=2, column=col, value=tf[key])
            c.data_type = "s"
        for col, w in ((1, 14), (2, 16), (3, 30), (4, 32)):
            ws.column_dimensions[get_column_letter(col)].width = w
        if include_fundamentals:
            _write_fundamentals_block(ws)

    # Manifest sheet (auditable).
    man = wb.create_sheet("Manifest")
    for r in _manifest_rows(tickers, depth, include_fundamentals):
        man.append(r)
    fg._style_header(man, 5)
    for col, w in ((1, 22), (2, 8), (3, 16), (4, 44), (5, 10)):
        man.column_dimensions[get_column_letter(col)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    # Reuse the proven cleanup that strips empty cached values and marks the
    # single-FDS spill cells as dynamic arrays.
    return fg._strip_empty_formula_values(bio.getvalue())


def build_weekly_templates_batched(
    tickers: List[str],
    depth: int = DEPTH,
    batch_size: int = BATCH_SIZE,
    include_fundamentals: bool = True,
) -> List[Tuple[str, bytes]]:
    """Split into chunks of ``batch_size`` and build one standalone workbook per
    chunk (each carries its own HSI sheet AND, when enabled, its own per-ticker
    fundamentals block). Returns [(filename, bytes), ...]."""
    chunks = fg._chunk(list(tickers), batch_size)
    total = len(chunks)
    width = max(2, len(str(total)))
    out: List[Tuple[str, bytes]] = []
    for idx, chunk in enumerate(chunks, start=1):
        first, last = chunk[0], chunk[-1]
        note = f"Batch {idx} of {total} — tickers {first}..{last} ({len(chunk)} names)"
        data = build_weekly_template(chunk, depth=depth, batch_note=note,
                                     include_fundamentals=include_fundamentals)
        fname = f"weekly_template_batch_{idx:0{width}d}_of_{total:0{width}d}.xlsx"
        out.append((fname, data))
    return out


def zip_templates(files: List[Tuple[str, bytes]]) -> bytes:
    return fg.zip_workbooks(files)
